"""
Excel 解析器
============

提供 ExcelParser 类和便捷函数，将 Excel 文件解析为结构化的 ExcelResult。

用法：
    # 便捷函数
    from .. import parse_excel, parse_directory

    result = parse_excel("/path/to/file.xlsx")
    results = parse_directory("/path/to/dir")

    # 类实例
    from .. import ExcelParser

    parser = ExcelParser(output_dir="./output", signing_detection=True)
    result = parser.parse("/path/to/file.xlsx")
"""

import os
import logging
from pathlib import Path
from typing import Optional, Union

import openpyxl

from .config import ENABLE_SIGNING_DETECTION, INCLUDE_HIDDEN
from .models import SheetContent, ExcelResult
from .converter import sheet_to_markdown

logger = logging.getLogger(__name__)


# 最小有效 styles.xml：非 Excel 生成的 xlsx 样式表损坏时回退使用。
# 关键点：cellXfs 需要提供足够多的 <xf> 占位，因为 sheet 里每个 cell
# 和列维度都引用了一个 style_id 索引（指向 cellXfs 数组），占位数不够
# 会触发 IndexError: list index out of range。
# 业务 Excel 文件实际样式数通常 < 500，预留 2 倍冗余。
_MAX_STYLE_ID_FALLBACK = 1000
_PADDING_XFS = "<xf/>" * _MAX_STYLE_ID_FALLBACK
_MIN_STYLES_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
    '<borders count="1"><border/></borders>'
    '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
    '<cellStyleXfs count="1"><xf/></cellStyleXfs>'
    f'<cellXfs count="{_MAX_STYLE_ID_FALLBACK}">{_PADDING_XFS}</cellXfs>'
    '</styleSheet>'
).encode("utf-8")


def _load_workbook_safe(file_path: str, data_only: bool, tmp_dir: Optional[str] = None):
    """加载 workbook；若样式表解析失败，则用剥离样式的临时副本重试。

    捕获的异常类型：
    - TypeError / ValueError / AttributeError：样式表对象类型不匹配
    - IndexError：cell 或列维度引用的 style_id 越界（cellXfs 数组不够长）

    Args:
        tmp_dir: 临时副本存放目录，None 时用系统默认（%TEMP% / /tmp）。
                 由上层 ExcelReader 从 PdfDocParserConfig.upload_tmp_dir 注入。
    """
    import zipfile
    import tempfile
    try:
        return openpyxl.load_workbook(file_path, data_only=data_only)
    except (TypeError, ValueError, AttributeError, IndexError) as e:
        msg = str(e)
        # IndexError 通常无 Fill/style 关键字，但 traceback 涉及 _cell_styles
        if not any(kw in msg for kw in ("Fill", "style", "styles", "PatternFill")):
            # 仅当 IndexError 来自样式相关栈帧时才回退
            if not isinstance(e, IndexError):
                raise
    logger.warning("openpyxl 样式表解析失败，剥离 styles.xml 后重试: %s", file_path)
    # 临时副本命名：{原文件名前缀}_fallback_{日期}_{时分秒毫秒}_{随机数}.xlsx
    # 便于排查"哪个 xlsx 触发了样式表降级"
    import datetime
    import random
    orig_stem = os.path.splitext(os.path.basename(file_path))[0]
    safe_stem = "".join(c if c.isalnum() or c in "_-" else "_" for c in orig_stem)[:40]
    now = datetime.datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S") + f"{now.microsecond // 1000:03d}"
    rnd = random.randint(0, 999)
    tmp_name = f"{safe_stem}_fallback_{ts}_{rnd:03d}.xlsx"
    if tmp_dir:
        tmp_path = os.path.join(tmp_dir, tmp_name)
    else:
        import tempfile as _tf
        tmp_path = os.path.join(_tf.gettempdir(), tmp_name)
    try:
        with zipfile.ZipFile(file_path, "r") as src, \
             zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.namelist():
                data = src.read(item)
                if item == "xl/styles.xml":
                    data = _MIN_STYLES_XML
                dst.writestr(item, data)
        return openpyxl.load_workbook(tmp_path, data_only=data_only)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


class ExcelParser:
    """
    Excel 文档解析器。

    将 Excel 文件解析为结构化结果，保留合并单元格、货币格式、
    分组表头展平、表单字段、签章信息等。

    Args:
        output_dir: 输出目录，传值则保存解析结果到磁盘，传 None 则仅返回内存结果
        signing_detection: 是否启用签章行检测，True=签章行剥离为元数据，False=当作普通数据行
        include_hidden: 是否包含隐藏行列，True=隐藏行列也输出，False=跳过隐藏行列
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        signing_detection: bool = ENABLE_SIGNING_DETECTION,
        include_hidden: bool = INCLUDE_HIDDEN,
        tmp_dir: Optional[str] = None,
    ):
        self.output_dir = output_dir
        self.signing_detection = signing_detection
        self.include_hidden = include_hidden
        self.tmp_dir = tmp_dir

    def parse(self, file_path: str, display_name: str | None = None) -> ExcelResult:
        """
        解析单个 Excel 文件。

        Args:
            file_path: Excel 文件路径（.xlsx / .xls）
            display_name: 在 Markdown 中显示的源文件名，None 时使用 file_path

        Returns:
            ExcelResult 解析结果
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx / .xls")

        # 动态设置签章检测开关和隐藏行列开关
        from . import config as cfg
        original_signing = cfg.ENABLE_SIGNING_DETECTION
        original_hidden = cfg.INCLUDE_HIDDEN
        cfg.ENABLE_SIGNING_DETECTION = self.signing_detection
        cfg.INCLUDE_HIDDEN = self.include_hidden

        wb = None
        wb_formula = None
        try:
            wb = _load_workbook_safe(file_path, data_only=True, tmp_dir=self.tmp_dir)
            wb_formula = _load_workbook_safe(file_path, data_only=False, tmp_dir=self.tmp_dir)

            # 构建完整 Markdown
            md_parts = []
            md_parts.append("# Excel 数据表")
            md_parts.append("")
            md_parts.append(f"> 源文件: `{display_name or file_path}`")
            md_parts.append(f"> 包含 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
            md_parts.append("")
            md_parts.append("## 目录")
            md_parts.append("")
            for idx, name in enumerate(wb.sheetnames, 1):
                md_parts.append(f"{idx}. [{name}](#{name.replace(' ', '-')})")
            md_parts.append("---")
            md_parts.append("")

            # 逐 Sheet 解析（sheet_to_markdown 正向返回 Markdown 和元数据）
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_f = wb_formula[sheet_name]
                sheet_md, sc = sheet_to_markdown(
                    ws, sheet_title=sheet_name,
                    ws_formula=ws_f, include_hidden=self.include_hidden,
                )
                md_parts.append(sheet_md)
                sheets.append(sc)

            full_markdown = "\n".join(md_parts)

            # 纯文本（去掉 Markdown 语法标记）
            full_text = _markdown_to_plain(full_markdown)

            # 元数据由 sheet_to_markdown 正向产出，无需反向解析

            result = ExcelResult(
                file_path=str(path),
                file_name=path.name,
                total_sheets=len(sheets),
                sheets=sheets,
                full_text=full_text,
                markdown_text=full_markdown,
                metadata={
                    "signing_detection": self.signing_detection,
                    "include_hidden": self.include_hidden,
                    "sheet_names": wb.sheetnames,
                },
            )

            # 保存到磁盘
            if self.output_dir:
                from .formatter import save_markdown, save_text
                out_dir = Path(self.output_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                stem = path.stem
                save_markdown(result, str(out_dir / f"{stem}.md"))
                save_text(result, str(out_dir / f"{stem}.txt"))

            return result

        finally:
            # 关闭工作簿，释放资源
            if wb is not None:
                wb.close()
            if wb_formula is not None:
                wb_formula.close()
            # 恢复原始配置
            cfg.ENABLE_SIGNING_DETECTION = original_signing
            cfg.INCLUDE_HIDDEN = original_hidden

    def parse_directory(self, dir_path: str, recursive: bool = False) -> list[ExcelResult]:
        """
        批量解析目录下的所有 Excel 文件。

        Args:
            dir_path: 目录路径
            recursive: 是否递归子目录

        Returns:
            ExcelResult 列表
        """
        dp = Path(dir_path)
        if not dp.is_dir():
            raise NotADirectoryError(f"不是目录: {dir_path}")

        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        files = list(dp.glob(pattern))
        if not recursive:
            files += list(dp.glob("*.xls"))
        else:
            files += list(dp.glob("**/*.xls"))

        # 排除临时文件
        files = [f for f in files if not f.name.startswith("~$")]

        results = []
        for f in sorted(files):
            try:
                logger.info(f"解析: {f.name}")
                result = self.parse(str(f))
                results.append(result)
            except Exception as e:
                logger.error(f"解析失败 [{f.name}]: {e}")

        return results


# ============================================================
# 内部辅助
# ============================================================

def _markdown_to_plain(md_text: str) -> str:
    """将 Markdown 转为纯文本（去掉表格竖线等格式标记）"""
    lines = md_text.split("\n")
    plain_lines = []
    for line in lines:
        stripped = line.strip()
        # 跳过分隔行
        if stripped.startswith("|") and all(c in "|- " for c in stripped):
            continue
        # 去掉表格竖线
        if stripped.startswith("|"):
            cells = [c.strip() for c in stripped.split("|") if c.strip()]
            plain_lines.append(" | ".join(cells))
        else:
            plain_lines.append(line)
    return "\n".join(plain_lines)


# ============================================================
# 便捷函数
# ============================================================

def parse_excel(
    file_path: str,
    output_dir: Optional[str] = None,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
    include_hidden: bool = INCLUDE_HIDDEN,
    tmp_dir: Optional[str] = None,
    display_name: Optional[str] = None,
) -> ExcelResult:
    """
    一键解析 Excel 文档。

    Args:
        file_path: Excel 文件路径（.xlsx / .xls）
        output_dir: 输出目录，传值则保存解析结果到磁盘，传 None 则仅返回内存结果
        signing_detection: 是否启用签章行检测
        include_hidden: 是否包含隐藏行列
        tmp_dir: 样式表降级副本的临时目录，None 时用系统默认（%TEMP% / /tmp）。
                 由上层 ExcelReader 从 PdfDocParserConfig.upload_tmp_dir 注入。

    Returns:
        ExcelResult 解析结果
    """
    parser = ExcelParser(
        output_dir=output_dir,
        signing_detection=signing_detection,
        include_hidden=include_hidden,
        tmp_dir=tmp_dir,
    )
    return parser.parse(file_path, display_name=display_name)


def parse_directory(
    dir_path: str,
    output_dir: Optional[str] = None,
    recursive: bool = False,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
    include_hidden: bool = INCLUDE_HIDDEN,
) -> list[ExcelResult]:
    """
    一键批量解析目录下的所有 Excel 文件。

    Args:
        dir_path: 目录路径
        output_dir: 输出目录
        recursive: 是否递归子目录
        signing_detection: 是否启用签章行检测
        include_hidden: 是否包含隐藏行列

    Returns:
        ExcelResult 列表
    """
    parser = ExcelParser(
        output_dir=output_dir,
        signing_detection=signing_detection,
        include_hidden=include_hidden,
    )
    return parser.parse_directory(dir_path, recursive=recursive)
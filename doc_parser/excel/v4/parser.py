"""
Excel V4 原始单元格 JSON 解析器
================================

提供 ExcelParser 类和便捷函数，将 Excel 文件解析为逐单元格的原始 JSON 数据。

用法：
    from . import parse_excel

    result = parse_excel("/path/to/file.xlsx")
    print(result.to_json())
"""

import logging
from pathlib import Path
from typing import Optional

import openpyxl

from .config import INCLUDE_HIDDEN, INCLUDE_EMPTY_CELLS
from .models import SheetRaw, ExcelRaw
from .converter import sheet_to_raw

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Excel 原始单元格 JSON 解析器。

    将 Excel 文件解析为逐单元格数据，保留合并单元格信息和格式化值。

    Args:
        output_dir: 输出目录，传值则保存 JSON 到磁盘
        include_hidden: 是否包含隐藏行列
        include_empty: 是否包含空单元格
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        include_hidden: bool = INCLUDE_HIDDEN,
        include_empty: bool = INCLUDE_EMPTY_CELLS,
    ):
        self.output_dir = output_dir
        self.include_hidden = include_hidden
        self.include_empty = include_empty

    def parse(self, file_path: str) -> ExcelRaw:
        """
        解析单个 Excel 文件。

        Args:
            file_path: Excel 文件路径（.xlsx / .xls）

        Returns:
            ExcelRaw 解析结果
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx / .xls")

        # 动态设置配置
        from . import config as cfg
        original_hidden = cfg.INCLUDE_HIDDEN
        original_empty = cfg.INCLUDE_EMPTY_CELLS
        cfg.INCLUDE_HIDDEN = self.include_hidden
        cfg.INCLUDE_EMPTY_CELLS = self.include_empty

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            wb_formula = openpyxl.load_workbook(file_path, data_only=False)

            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_f = wb_formula[sheet_name]
                data = sheet_to_raw(
                    ws, sheet_title=sheet_name,
                    ws_formula=ws_f,
                    include_hidden=self.include_hidden,
                    include_empty=self.include_empty,
                )
                sc = SheetRaw(
                    sheet_name=data["sheet_name"],
                    max_row=data["max_row"],
                    max_col=data["max_col"],
                    merged_cells=data["merged_cells"],
                    cells=data["cells"],
                )
                sheets.append(sc)

            result = ExcelRaw(
                file_path=str(path),
                file_name=path.name,
                total_sheets=len(sheets),
                sheets=sheets,
                metadata={
                    "include_hidden": self.include_hidden,
                    "include_empty": self.include_empty,
                    "sheet_names": wb.sheetnames,
                },
            )

            # 保存到磁盘
            if self.output_dir:
                from .formatter import save_json
                out_dir = Path(self.output_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                stem = path.stem
                save_json(result, str(out_dir / f"{stem}.json"))

            return result

        finally:
            cfg.INCLUDE_HIDDEN = original_hidden
            cfg.INCLUDE_EMPTY_CELLS = original_empty

    def parse_directory(self, dir_path: str, recursive: bool = False) -> list[ExcelRaw]:
        """
        批量解析目录下的所有 Excel 文件。

        Args:
            dir_path: 目录路径
            recursive: 是否递归子目录

        Returns:
            ExcelRaw 列表
        """
        dp = Path(dir_path)
        if not dp.is_dir():
            raise NotADirectoryError(f"不是目录: {dir_path}")

        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        files = list(dp.glob(pattern))
        if not recursive:
            files += list(dp.glob("*.xls"))
        else:
            files += list(dp.glob("**/*.xls"))

        files = [f for f in files if not f.name.startswith("~$")]

        results = []
        for f in sorted(files):
            try:
                logger.info(f"解析: {f.name}")
                result = self.parse(str(f))
                results.append(result)
            except Exception as e:
                logger.error(f"解析失败 [{f.name}]: {e}")

        return results


# ============================================================
# 便捷函数
# ============================================================

def parse_excel(
    file_path: str,
    output_dir: Optional[str] = None,
    include_hidden: bool = INCLUDE_HIDDEN,
    include_empty: bool = INCLUDE_EMPTY_CELLS,
) -> ExcelRaw:
    """
    一键解析 Excel 文档为原始单元格 JSON。

    Args:
        file_path: Excel 文件路径
        output_dir: 输出目录
        include_hidden: 是否包含隐藏行列
        include_empty: 是否包含空单元格

    Returns:
        ExcelRaw 解析结果
    """
    parser = ExcelParser(
        output_dir=output_dir,
        include_hidden=include_hidden,
        include_empty=include_empty,
    )
    return parser.parse(file_path)


def parse_directory(
    dir_path: str,
    output_dir: Optional[str] = None,
    recursive: bool = False,
    include_hidden: bool = INCLUDE_HIDDEN,
    include_empty: bool = INCLUDE_EMPTY_CELLS,
) -> list[ExcelRaw]:
    """
    一键批量解析目录下的所有 Excel 文件。

    Args:
        dir_path: 目录路径
        output_dir: 输出目录
        recursive: 是否递归子目录
        include_hidden: 是否包含隐藏行列
        include_empty: 是否包含空单元格

    Returns:
        ExcelRaw 列表
    """
    parser = ExcelParser(
        output_dir=output_dir,
        include_hidden=include_hidden,
        include_empty=include_empty,
    )
    return parser.parse_directory(dir_path, recursive=recursive)

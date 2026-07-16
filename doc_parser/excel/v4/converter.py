"""
Excel 原始单元格采集与格式化
============================

将工作表转换为逐单元格的原始数据格式：
- 每个单元格以 "A1": value 形式输出
- 合并单元格区域内非起点位置输出 null
- 值经过格式化处理（货币、百分比、日期等）
"""

from typing import Any, Optional, List, Dict, Tuple

from openpyxl.utils import get_column_letter

from .config import INCLUDE_HIDDEN, INCLUDE_EMPTY_CELLS


# ============================================================
# 隐藏行列检测
# ============================================================

def _has_active_autofilter(ws) -> bool:
    """检测 sheet 是否存在生效中的 AutoFilter 筛选条件。

    openpyxl 的 AutoFilter 陷阱：Excel 开启数据筛选后，不满足筛选条件的行会被
    标记为 row_dimensions[r].hidden=True。这种"隐藏"与用户手动右键隐藏的行
    在 hidden 属性上无法区分，但语义完全不同：
    - 手动隐藏：用户有意排除的辅助行/计算行，可过滤
    - AutoFilter 隐藏：用户在 Excel 里的临时查看筛选，数据本身完整，入库必须读取

    Excel 有两种筛选机制，均会引发上述 hidden 标记，需同时检测：
    1. sheet 级 AutoFilter：选中区域→数据→筛选，对应 ws.auto_filter
    2. Table 级筛选：Ctrl+T 转成表格后表头筛选，对应 ws.tables[*].autoFilter
       （注意 openpyxl 命名不一致：sheet 级是 auto_filter，Table 级是 autoFilter）

    判定依据：筛选区域非空（ref）且实际勾选了筛选值（filterColumn 非空）。
    仅 ref 没有 filterColumn 时是"未应用筛选条件"，不会产生隐藏行，无需特殊处理。
    """
    # 1. sheet 级 AutoFilter
    af = getattr(ws, "auto_filter", None)
    if af is not None and af.ref and af.filterColumn:
        return True
    # 2. Table 级筛选（ListObject）
    tables = getattr(ws, "tables", None)
    if tables:
        for tbl in tables.values():
            taf = getattr(tbl, "autoFilter", None)
            if taf is not None and taf.filterColumn:
                return True
    return False


def _get_visible_rows_cols(ws) -> Tuple[List[int], List[int]]:
    """获取可见（非隐藏）的行号和列号列表。

    AutoFilter 修正：当 sheet 存在生效的 AutoFilter 筛选条件时，忽略所有行的
    hidden 标记——因为这些隐藏是筛选产生的，数据本身完整，入库场景必须全部读取。
    否则正常过滤手动隐藏的行列。
    """
    # AutoFilter 场景：筛选产生的隐藏行不算真隐藏，全部读取
    if _has_active_autofilter(ws):
        visible_rows = list(range(1, (ws.max_row or 0) + 1))
        visible_cols = list(range(1, (ws.max_column or 0) + 1))
        return visible_rows, visible_cols

    hidden_rows = set()
    for r, dim in ws.row_dimensions.items():
        if dim.hidden:
            hidden_rows.add(r)

    hidden_cols = set()
    for col_key, dim in ws.column_dimensions.items():
        if dim.hidden:
            if isinstance(col_key, int):
                hidden_cols.add(col_key)
            else:
                from openpyxl.utils import column_index_from_string
                hidden_cols.add(column_index_from_string(col_key))

    visible_rows = sorted(r for r in range(1, (ws.max_row or 0) + 1) if r not in hidden_rows)
    visible_cols = sorted(c for c in range(1, (ws.max_column or 0) + 1) if c not in hidden_cols)
    return visible_rows, visible_cols


# ============================================================
# 合并单元格信息采集
# ============================================================

def _collect_merged_ranges(ws) -> Tuple[Dict[Tuple[int, int], Tuple[int, int]], List[str]]:
    """
    采集合并单元格信息。

    Returns:
        origin_map: {(row, col): (min_row, min_col)} 每个合并区域内单元格指向其起点
        range_strs: ["A1:H1", ...] 合并区域的引用字符串列表
    """
    origin_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    range_strs: List[str] = []

    for merged_range in ws.merged_cells.ranges:
        min_r, max_r = merged_range.min_row, merged_range.max_row
        min_c, max_c = merged_range.min_col, merged_range.max_col
        ref = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
        range_strs.append(ref)

        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue
                origin_map[(r, c)] = (min_r, min_c)

    return origin_map, range_strs


# ============================================================
# 单元格值格式化
# ============================================================

def _format_value(value: Any, number_format: str) -> str:
    """
    格式化单元格值，使其对 LLM 可读。

    支持：货币、百分比、千分位、日期、布尔、文本
    """
    if value is None:
        return ""

    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>").strip()

    if isinstance(value, bool):
        return "是" if value else "否"

    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)

    if isinstance(value, (int, float)):
        fmt = number_format or ""

        # 货币格式
        for symbol, prefix in [("¥", "¥"), ("￥", "¥"), ("$", "$"), ("€", "€")]:
            if symbol in fmt:
                return f"{prefix}{value:,.2f}"

        # 百分比
        if "%" in fmt:
            pct = value * 100
            if pct == int(pct):
                return f"{int(pct)}%"
            return f"{pct:.2f}%"

        # 千分位
        if "#,##0" in fmt or "#,#0" in fmt:
            if ".00" in fmt or ".0" in fmt:
                return f"{value:,.2f}"
            if isinstance(value, float) and value != int(value):
                return f"{value:,.2f}"
            return f"{int(value):,}"

        # 通用数值
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(round(value, 10))

        return str(value)

    return str(value)


# ============================================================
# 主处理函数
# ============================================================

def sheet_to_raw(ws, sheet_title: Optional[str] = None,
                 ws_formula=None,
                 include_hidden: bool = False,
                 include_empty: bool = True) -> dict:
    """
    将单个 Sheet 转换为原始单元格 JSON dict。

    Args:
        ws: 工作表对象（data_only=True 加载）
        sheet_title: 工作表标题，默认使用 ws.title
        ws_formula: 以 data_only=False 加载的工作表，用于提取公式文本
        include_hidden: 是否包含隐藏行列
        include_empty: 是否包含空单元格（为 null）

    Returns:
        dict: {sheet_name, max_row, max_col, merged_cells, cells}
    """
    title = sheet_title or ws.title

    if ws.max_row is None or ws.max_column is None or ws.max_row == 0:
        return {
            "sheet_name": title,
            "max_row": 0,
            "max_col": 0,
            "merged_cells": [],
            "cells": {},
        }

    # 获取可见行列
    if include_hidden:
        visible_rows = list(range(1, ws.max_row + 1))
        visible_cols = list(range(1, ws.max_column + 1))
    else:
        visible_rows, visible_cols = _get_visible_rows_cols(ws)

    # 合并单元格信息
    origin_map, range_strs = _collect_merged_ranges(ws)

    # 采集公式：从 formula 工作表中获取公式文本
    formula_map: Dict[Tuple[int, int], str] = {}
    if ws_formula is not None:
        for row_idx in visible_rows:
            for col_idx in visible_cols:
                fcell = ws_formula.cell(row=row_idx, column=col_idx)
                if fcell.value and isinstance(fcell.value, str) and fcell.value.startswith("="):
                    formula_map[(row_idx, col_idx)] = fcell.value

    # 逐单元格采集
    cells: Dict[str, Any] = {}

    for row_idx in visible_rows:
        for col_idx in visible_cols:
            ref = f"{get_column_letter(col_idx)}{row_idx}"

            # 合并区域内非起点 → null
            if (row_idx, col_idx) in origin_map:
                if include_empty:
                    cells[ref] = None
                continue

            cell = ws.cell(row=row_idx, column=col_idx)

            # 公式单元格：直接输出公式文本
            if (row_idx, col_idx) in formula_map:
                cells[ref] = formula_map[(row_idx, col_idx)]
                continue

            value = cell.value

            if value is None or (isinstance(value, str) and value.strip() == ""):
                if include_empty:
                    cells[ref] = None
                continue

            # 格式化值
            number_format = cell.number_format or "General"
            formatted = _format_value(value, number_format)
            cells[ref] = formatted

    return {
        "sheet_name": title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_cells": range_strs,
        "cells": cells,
    }
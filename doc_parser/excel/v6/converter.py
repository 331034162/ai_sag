"""
Excel 工作表转 CSV 格式
========================

将每个 Sheet 转为标准 CSV 文本：
- 逗号分隔，双引号包裹含逗号/换行/双引号的字段
- 合并单元格：
  - 行方向合并（同列跨多行）：向下填充起点值，保证每行语义完整（分组不丢失）
  - 列方向合并（同行跨多列）：起点填值，非起点留空（保持现状）
  - 标题行/表单行/签章行/分组表头行的纵向合并不做向下填充（避免把标题/表单值污染到数据行）
- 值经过格式化处理（货币、百分比、日期等）
- 公式单元格输出计算值

合并处理策略综合了 V2/V3 的能力：
- V3 的 apply_merge_logic：行方向合并向下填充（保留分组上下文）
- V2 的结构识别：_is_title_row / _is_form_row / _is_signing_row / _is_group_header
  （用于判断行角色，避免对标题/表单/签章行误填充）
"""

import csv
import io
import re
from typing import Any, Optional, List, Dict, Tuple

from .config import (
    INCLUDE_HIDDEN,
    INCLUDE_EMPTY_CELLS,
    SIGNING_KEYWORDS,
)


# ============================================================
# 隐藏行列检测
# ============================================================

def _has_active_autofilter(ws) -> bool:
    """检测 sheet 是否存在生效中的 AutoFilter 筛选条件。

    openpyxl 的 AutoFilter 陷阱：Excel 开启数据筛选后，不满足筛选条件的行会被
    标记为 row_dimensions[r].hidden=True。这种"隐藏"与用户手动右键隐藏的行
    在 hidden 属性上无法区分，但语义完全不同：
    - 手动隐藏：用户有意排除的辅助行/计算行，可过滤
    - AutoFilter 隐藏：用户在 Excel 里的临时查看筛选，数据本身完整，入库必须读取

    Excel 有两种筛选机制，均会引发上述 hidden 标记，需同时检测：
    1. sheet 级 AutoFilter：选中区域→数据→筛选，对应 ws.auto_filter
    2. Table 级筛选：Ctrl+T 转成表格后表头筛选，对应 ws.tables[*].autoFilter
       （注意 openpyxl 命名不一致：sheet 级是 auto_filter，Table 级是 autoFilter）

    判定依据：筛选区域非空（ref）且实际勾选了筛选值（filterColumn 非空）。
    仅 ref 没有 filterColumn 时是"未应用筛选条件"，不会产生隐藏行，无需特殊处理。
    """
    # 1. sheet 级 AutoFilter
    af = getattr(ws, "auto_filter", None)
    if af is not None and af.ref and af.filterColumn:
        return True
    # 2. Table 级筛选（ListObject）
    tables = getattr(ws, "tables", None)
    if tables:
        for tbl in tables.values():
            taf = getattr(tbl, "autoFilter", None)
            if taf is not None and taf.filterColumn:
                return True
    return False


def _get_visible_rows_cols(ws) -> Tuple[List[int], List[int]]:
    """获取可见（非隐藏）的行号和列号列表。

    AutoFilter 修正：当 sheet 存在生效的 AutoFilter 筛选条件时，忽略所有行的
    hidden 标记——因为这些隐藏是筛选产生的，数据本身完整，入库场景必须全部读取。
    否则正常过滤手动隐藏的行列。
    """
    # AutoFilter 场景：筛选产生的隐藏行不算真隐藏，全部读取
    if _has_active_autofilter(ws):
        visible_rows = list(range(1, (ws.max_row or 0) + 1))
        visible_cols = list(range(1, (ws.max_column or 0) + 1))
        return visible_rows, visible_cols

    hidden_rows = set()
    for r, dim in ws.row_dimensions.items():
        if dim.hidden:
            hidden_rows.add(r)

    hidden_cols = set()
    for col_key, dim in ws.column_dimensions.items():
        if dim.hidden:
            if isinstance(col_key, int):
                hidden_cols.add(col_key)
            else:
                from openpyxl.utils import column_index_from_string
                hidden_cols.add(column_index_from_string(col_key))

    visible_rows = sorted(r for r in range(1, (ws.max_row or 0) + 1) if r not in hidden_rows)
    visible_cols = sorted(c for c in range(1, (ws.max_column or 0) + 1) if c not in hidden_cols)
    return visible_rows, visible_cols


# ============================================================
# 合并单元格信息采集
# ============================================================

def _collect_merged_ranges(ws) -> Tuple[Dict[Tuple[int, int], Tuple[int, int, int, int]], List[Tuple[int, int, int, int]]]:
    """
    采集合并单元格信息。

    Returns:
        origin_map: {(row, col): (min_row, min_col, max_row, max_col)}
                    每个合并区域内的单元格指向其所在合并区域的完整边界
        merge_list: [(min_row, min_col, max_row, max_col), ...]
                    所有合并区域列表，用于遍历
    """
    origin_map: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    merge_list: List[Tuple[int, int, int, int]] = []

    for merged_range in ws.merged_cells.ranges:
        min_r, max_r = merged_range.min_row, merged_range.max_row
        min_c, max_c = merged_range.min_col, merged_range.max_col
        bounds = (min_r, min_c, max_r, max_c)
        merge_list.append(bounds)

        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue
                origin_map[(r, c)] = bounds

    return origin_map, merge_list


# ============================================================
# 结构识别（移植自 V2，用于判断行角色，避免误填充）
# ============================================================

def _is_title_row(row_data: list, merge_origins_in_row: List[Tuple[int, int, int, int]], num_cols: int) -> bool:
    """标题行：单个值横跨大部分列（横向合并且 col_span >= 50% 列数）。

    例: A1:H1="供应链原材料采购入库单" → 标题行
    """
    non_empty = [v for v in row_data if v.strip()]
    if len(non_empty) != 1:
        return False
    for (min_r, min_c, max_r, max_c) in merge_origins_in_row:
        col_span = max_c - min_c + 1
        if col_span >= num_cols * 0.5:
            return True
    return False


def _is_form_row(row_data: list, merge_origins_in_row: List[Tuple[int, int, int, int]]) -> bool:
    """表单字段行：横向合并起点含 key：value 形式。

    例: A2:B2="供应商：武汉XX科技有限公司" → 表单行
    """
    for (min_r, min_c, max_r, max_c) in merge_origins_in_row:
        if max_c > min_c:  # 横向合并
            # 取起点单元格的值（row_data 中第一个非空即起点值）
            idx = min_c - 1
            if 0 <= idx < len(row_data):
                val = row_data[idx]
                if val and ("：" in val or ":" in val):
                    return True
    return False


def _is_signing_row(row_data: list) -> bool:
    """签章行检测。

    严格匹配：签字/签章/公章等关键词直接匹配。
    审批关键词：仅匹配标签形式（如"部门经理审批："），不匹配描述文本中的嵌入词。
    """
    for v in row_data:
        v = v.strip()
        if not v:
            continue
        if any(kw in v for kw in SIGNING_KEYWORDS):
            return True
        if '审批' in v:
            if re.search(r'审批[：:]', v):
                return True
            if len(v) <= 10 and v.endswith('审批'):
                return True
    return False


def _has_col_merges(merge_origins_in_row: List[Tuple[int, int, int, int]]) -> bool:
    """判断该行是否有列合并（横向合并）起点"""
    for (min_r, min_c, max_r, max_c) in merge_origins_in_row:
        if max_c > min_c:
            return True
    return False


def _is_group_header(row_pos: int, rows: List[Tuple[int, list, List[Tuple[int, int, int, int]]]],
                     num_cols: int) -> bool:
    """判断是否为段落分组表头行。

    条件：
    1. 该行有列合并
    2. 后续紧跟着一行叶子表头（无列合并、非空值够多）

    例: 供应商资质信息表第1行 B1:E1="基础工商信息" + 第2行 "公司名称|统一社会信用代码|..."
    → 第1行是分组表头
    """
    if row_pos >= len(rows):
        return False
    _, _, merge_origins = rows[row_pos]
    if not _has_col_merges(merge_origins):
        return False

    for j in range(row_pos + 1, min(row_pos + 5, len(rows))):
        _, next_data, next_merges = rows[j]
        if all(v.strip() == "" for v in next_data):
            continue
        if _has_col_merges(next_merges):
            return True
        non_empty = sum(1 for v in next_data if v.strip())
        if non_empty >= max(num_cols * 0.3, 2):
            return True
        break
    return False


# ============================================================
# 角色前缀映射（结构识别开启时输出到 CSV 行首，TableSplitter 据此定位真表头/跳过非数据行）
# ============================================================

_ROLE_PREFIX = {
    "title": "#TITLE#",
    "form": "#FORM#",
    "signing": "#SIGNING#",
    "group_header": "#GROUP_HEADER#",
}


# ============================================================
# 向下填充决策
# ============================================================

def _build_row_roles(rows: List[Tuple[int, list, List[Tuple[int, int, int, int]]]],
                     num_cols: int) -> Dict[int, str]:
    """为每行判定角色：title / form / signing / group_header / data。

    Returns:
        {row_idx: role} 未匹配的行不写入（默认 data）
    """
    roles: Dict[int, str] = {}
    for i, (row_idx, row_data, merge_origins) in enumerate(rows):
        if _is_title_row(row_data, merge_origins, num_cols):
            roles[row_idx] = "title"
            continue
        # signing 优先于 form：签章行（如"申请人签字："）含"："会被 form 误判，需先判签章
        if _is_signing_row(row_data):
            roles[row_idx] = "signing"
            continue
        if _is_form_row(row_data, merge_origins):
            roles[row_idx] = "form"
            continue
        if _is_group_header(i, rows, num_cols):
            roles[row_idx] = "group_header"
            continue
    return roles


# ============================================================
# 单元格值格式化
# ============================================================

def _format_value(value: Any, number_format: str) -> str:
    """格式化单元格值，使其对 LLM 可读"""
    if value is None:
        return ""

    if isinstance(value, str):
        return value.replace("\r\n", " ").replace("\r", " ").replace("\n", " ").strip()

    if isinstance(value, bool):
        return "是" if value else "否"

    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)

    if isinstance(value, (int, float)):
        fmt = number_format or ""

        for symbol, prefix in [("¥", "¥"), ("￥", "¥"), ("$", "$"), ("€", "€")]:
            if symbol in fmt:
                return f"{prefix}{value:,.2f}"

        if "%" in fmt:
            pct = value * 100
            if pct == int(pct):
                return f"{int(pct)}%"
            return f"{pct:.2f}%"

        if "#,##0" in fmt or "#,#0" in fmt:
            if ".00" in fmt or ".0" in fmt:
                return f"{value:,.2f}"
            if isinstance(value, float) and value != int(value):
                return f"{value:,.2f}"
            return f"{int(value):,}"

        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(round(value, 10))

        return str(value)

    return str(value)


# ============================================================
# 主处理函数
# ============================================================

def sheet_to_csv(ws, sheet_title: Optional[str] = None,
                 include_hidden: bool = False,
                 include_empty: bool = True) -> dict:
    """
    将单个 Sheet 转换为 CSV 文本。

    合并单元格处理受两个开关控制（见 config.py）：
    - ENABLE_MERGE_FILL_DOWN：纵向合并（同列跨多行）是否向下填充起点值
    - ENABLE_STRUCTURE_DETECTION：是否识别标题/表单/签章/分组表头行（填充时保护这些行）

    开关组合：
    - 都关：原始行为，合并非起点一律留空
    - 只开填充：无脑向下填充纵向合并（有误填充标题列的风险）
    - 都开（默认）：智能填充，结构识别保护 + 纵向向下填充

    Args:
        ws: 工作表对象（data_only=True 加载）
        sheet_title: 工作表标题
        include_hidden: 是否包含隐藏行列
        include_empty: 是否包含空行

    Returns:
        dict: {sheet_name, csv_text, row_count, col_count}
    """
    title = sheet_title or ws.title

    if ws.max_row is None or ws.max_column is None or ws.max_row == 0:
        return {
            "sheet_name": title,
            "csv_text": "",
            "row_count": 0,
            "col_count": 0,
        }

    # 获取可见行列
    if include_hidden:
        visible_rows = list(range(1, ws.max_row + 1))
        visible_cols = list(range(1, ws.max_column + 1))
    else:
        visible_rows, visible_cols = _get_visible_rows_cols(ws)

    if not visible_rows or not visible_cols:
        return {
            "sheet_name": title,
            "csv_text": "",
            "row_count": 0,
            "col_count": len(visible_cols),
        }

    # 合并单元格信息
    origin_map, merge_list = _collect_merged_ranges(ws)
    num_cols = len(visible_cols)

    # 动态读取开关（运行时可改，与 parser.py 一致，避免 from import 快照失效）
    from . import config as cfg
    use_fill_down = cfg.ENABLE_MERGE_FILL_DOWN
    use_structure = cfg.ENABLE_STRUCTURE_DETECTION

    # 预读所有"普通单元格 + 合并起点"的格式化值（非起点不读，后续按规则决定）
    # 同时作为结构识别的原始数据和向下填充的起点值来源
    cell_cache: Dict[Tuple[int, int], str] = {}
    for row_idx in visible_rows:
        for col_idx in visible_cols:
            if (row_idx, col_idx) in origin_map:
                continue
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None or (isinstance(value, str) and value.strip() == ""):
                cell_cache[(row_idx, col_idx)] = ""
            else:
                number_format = cell.number_format or "General"
                cell_cache[(row_idx, col_idx)] = _format_value(value, number_format)

    # 构建行结构：(row_idx, raw_row_data, 该行的合并起点bounds列表)
    origins_by_row: Dict[int, List[Tuple[int, int, int, int]]] = {}
    for bounds in merge_list:
        origins_by_row.setdefault(bounds[0], []).append(bounds)

    rows_meta: List[Tuple[int, list, List[Tuple[int, int, int, int]]]] = []
    for row_idx in visible_rows:
        raw_row = [cell_cache.get((row_idx, c), "") for c in visible_cols]
        rows_meta.append((row_idx, raw_row, origins_by_row.get(row_idx, [])))

    # 行角色判定（仅当填充+结构识别都开启时才需要）
    roles: Dict[int, str] = {}
    if use_fill_down and use_structure:
        roles = _build_row_roles(rows_meta, num_cols)

    protected_roles = {"title", "form", "signing", "group_header"}

    # 按行输出 CSV（结构识别开启时给每行加角色前缀，便于 TableSplitter 定位真表头/跳过非数据行）
    lines: List[str] = []
    row_count = 0

    for row_idx in visible_rows:
        row_values = []
        has_value = False
        role = roles.get(row_idx)

        for col_idx in visible_cols:
            pos = (row_idx, col_idx)

            if pos not in origin_map:
                # 普通单元格或合并起点 → 直接取格式化值
                val = cell_cache.get(pos, "")
                row_values.append(val)
                if val:
                    has_value = True
                continue

            # 合并非起点单元格
            bounds = origin_map[pos]
            min_r, min_c, max_r, max_c = bounds

            # 判断是否需要向下填充（仅纵向合并的同列延续位置）
            # 纯纵向合并（max_c == min_c）才填充；块合并（跨行又跨列）不填充，
            # 因为块合并多为签章区域/标题块，填充会污染上下文
            should_fill = False
            if (use_fill_down and col_idx == min_c
                    and max_r > min_r and max_c == min_c):
                should_fill = True
                # 结构识别保护：标题/表单/签章/分组表头行不填充
                if use_structure and role in protected_roles:
                    should_fill = False

            if should_fill:
                val = cell_cache.get((min_r, min_c), "")
                row_values.append(val)
                if val:
                    has_value = True
            else:
                row_values.append("")

        # 跳过全空行
        if not has_value and not include_empty:
            continue

        # 生成标准 CSV 行文本（用临时 StringIO 复用 csv.writer 的转义逻辑）
        tmp = io.StringIO()
        csv.writer(tmp, quoting=csv.QUOTE_MINIMAL).writerow(row_values)
        row_csv = tmp.getvalue().rstrip("\r\n")

        # 结构识别开启时加角色前缀；data 行用 #DATA#，其余按角色映射
        # TableSplitter 据此：跳过 TITLE/FORM/SIGNING/GROUP_HEADER 行，
        # 取第一个 DATA 行作真表头（叶子表头），后续 DATA 行作数据行
        if use_structure:
            prefix = _ROLE_PREFIX.get(role, "#DATA#")
            lines.append(prefix + row_csv)
        else:
            lines.append(row_csv)

        if has_value:
            row_count += 1

    csv_text = "\n".join(lines)

    return {
        "sheet_name": title,
        "csv_text": csv_text,
        "row_count": row_count,
        "col_count": len(visible_cols),
    }
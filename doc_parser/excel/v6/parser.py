"""
Excel V6 CSV 格式解析器
========================

提供 ExcelParser 类和便捷函数，将 Excel 文件解析为 CSV 文本。

用法：
    from . import parse_excel

    result = parse_excel("/path/to/file.xlsx")
    print(result.to_csv_text())
"""

import logging
import os
from pathlib import Path
from typing import Optional

import openpyxl


# 最小有效 styles.xml：非 Excel 生成的 xlsx 样式表损坏时回退使用。
# 关键点：cellXfs 需要提供足够多的 <xf> 占位，因为 sheet 里每个 cell
# 和列维度都引用了一个 style_id 索引（指向 cellXfs 数组），占位数不够
# 会触发 IndexError: list index out of range。
# 业务 Excel 文件实际样式数通常 < 500，预留 2 倍冗余。
_MAX_STYLE_ID_FALLBACK = 1000
_PADDING_XFS = "<xf/>" * _MAX_STYLE_ID_FALLBACK
_MIN_STYLES_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
    '<borders count="1"><border/></borders>'
    '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
    '<cellStyleXfs count="1"><xf/></cellStyleXfs>'
    f'<cellXfs count="{_MAX_STYLE_ID_FALLBACK}">{_PADDING_XFS}</cellXfs>'
    '</styleSheet>'
).encode("utf-8")


def _load_workbook_safe(file_path: str, data_only: bool, tmp_dir: Optional[str] = None):
    """加载 workbook；若样式表解析失败，则用剥离样式的临时副本重试。

    捕获的异常类型：
    - TypeError / ValueError / AttributeError：样式表对象类型不匹配
      （如 openpyxl 3.1.5 遇到空 ``<fill></fill>`` 时抛出
      ``TypeError: Fill() takes no arguments``）
    - IndexError：cell 或列维度引用的 style_id 越界（cellXfs 数组不够长）

    Args:
        tmp_dir: 临时副本存放目录，None 时用系统默认（%TEMP% / /tmp）。
    """
    import zipfile
    import tempfile
    try:
        return openpyxl.load_workbook(file_path, data_only=data_only)
    except (TypeError, ValueError, AttributeError, IndexError) as e:
        msg = str(e)
        # IndexError 通常无 Fill/style 关键字，但 traceback 涉及 _cell_styles
        if not any(kw in msg for kw in ("Fill", "style", "styles", "PatternFill")):
            # 仅当 IndexError 来自样式相关栈帧时才回退
            if not isinstance(e, IndexError):
                raise
    logger.warning("openpyxl 样式表解析失败，剥离 styles.xml 后重试: %s", file_path)
    # 临时副本命名：{原文件名前缀}_fallback_{日期}_{时分秒毫秒}_{随机数}.xlsx
    import datetime
    import random
    orig_stem = os.path.splitext(os.path.basename(file_path))[0]
    safe_stem = "".join(c if c.isalnum() or c in "_-" else "_" for c in orig_stem)[:40]
    now = datetime.datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S") + f"{now.microsecond // 1000:03d}"
    rnd = random.randint(0, 999)
    tmp_name = f"{safe_stem}_fallback_{ts}_{rnd:03d}.xlsx"
    if tmp_dir:
        tmp_path = os.path.join(tmp_dir, tmp_name)
    else:
        import tempfile as _tf
        tmp_path = os.path.join(_tf.gettempdir(), tmp_name)
    try:
        with zipfile.ZipFile(file_path, "r") as src, \
             zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.namelist():
                data = src.read(item)
                if item == "xl/styles.xml":
                    data = _MIN_STYLES_XML
                dst.writestr(item, data)
        return openpyxl.load_workbook(tmp_path, data_only=data_only)
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


from .config import INCLUDE_HIDDEN, INCLUDE_EMPTY_CELLS
from .models import SheetCSV, ExcelCSV
from .converter import sheet_to_csv

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Excel CSV 格式解析器。

    将 Excel 文件解析为 CSV 文本，每个 Sheet 输出为逗号分隔的纯文本。

    Args:
        output_dir: 输出目录，传值则保存到磁盘
        include_hidden: 是否包含隐藏行列
        include_empty: 是否包含空行
        tmp_dir: 样式表降级副本的临时目录，None 时用系统默认（%TEMP% / /tmp）。
                 由上层 ExcelReader 从 PdfDocParserConfig.upload_tmp_dir 注入。
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        include_hidden: bool = INCLUDE_HIDDEN,
        include_empty: bool = INCLUDE_EMPTY_CELLS,
        tmp_dir: Optional[str] = None,
    ):
        self.output_dir = output_dir
        self.include_hidden = include_hidden
        self.include_empty = include_empty
        self.tmp_dir = tmp_dir

    def parse(self, file_path: str) -> ExcelCSV:
        """解析单个 Excel 文件为 CSV 文本"""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx / .xls")

        # 动态设置配置
        from . import config as cfg
        original_hidden = cfg.INCLUDE_HIDDEN
        original_empty = cfg.INCLUDE_EMPTY_CELLS
        cfg.INCLUDE_HIDDEN = self.include_hidden
        cfg.INCLUDE_EMPTY_CELLS = self.include_empty

        wb = None
        try:
            # data_only=True 获取公式计算值；样式表损坏时自动降级重试
            wb = _load_workbook_safe(file_path, data_only=True, tmp_dir=self.tmp_dir)

            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = sheet_to_csv(
                    ws, sheet_title=sheet_name,
                    include_hidden=self.include_hidden,
                    include_empty=self.include_empty,
                )
                sc = SheetCSV(
                    sheet_name=data["sheet_name"],
                    csv_text=data["csv_text"],
                    row_count=data["row_count"],
                    col_count=data["col_count"],
                )
                sheets.append(sc)

            result = ExcelCSV(
                file_path=str(path),
                file_name=path.name,
                total_sheets=len(sheets),
                sheets=sheets,
                metadata={
                    "include_hidden": self.include_hidden,
                    "include_empty": self.include_empty,
                    "sheet_names": wb.sheetnames,
                },
            )

            # 保存到磁盘
            if self.output_dir:
                from .formatter import save_csv_text
                out_dir = Path(self.output_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                save_csv_text(result, str(out_dir / f"{path.stem}.csv"))

            return result

        finally:
            if wb is not None:
                wb.close()
            cfg.INCLUDE_HIDDEN = original_hidden
            cfg.INCLUDE_EMPTY_CELLS = original_empty

    def parse_directory(self, dir_path: str, recursive: bool = False) -> list[ExcelCSV]:
        """批量解析目录下的所有 Excel 文件"""
        dp = Path(dir_path)
        if not dp.is_dir():
            raise NotADirectoryError(f"不是目录: {dir_path}")

        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        files = list(dp.glob(pattern))
        if not recursive:
            files += list(dp.glob("*.xls"))
        else:
            files += list(dp.glob("**/*.xls"))

        files = [f for f in files if not f.name.startswith("~$")]

        results = []
        for f in sorted(files):
            try:
                logger.info(f"解析: {f.name}")
                result = self.parse(str(f))
                results.append(result)
            except Exception as e:
                logger.error(f"解析失败 [{f.name}]: {e}")

        return results


# ============================================================
# 便捷函数
# ============================================================

def parse_excel(
    file_path: str,
    output_dir: Optional[str] = None,
    include_hidden: bool = INCLUDE_HIDDEN,
    include_empty: bool = INCLUDE_EMPTY_CELLS,
    tmp_dir: Optional[str] = None,
) -> ExcelCSV:
    """一键解析 Excel 文档为 CSV 文本"""
    parser = ExcelParser(
        output_dir=output_dir,
        include_hidden=include_hidden,
        include_empty=include_empty,
        tmp_dir=tmp_dir,
    )
    return parser.parse(file_path)


def parse_directory(
    dir_path: str,
    output_dir: Optional[str] = None,
    recursive: bool = False,
    include_hidden: bool = INCLUDE_HIDDEN,
    include_empty: bool = INCLUDE_EMPTY_CELLS,
) -> list[ExcelCSV]:
    """一键批量解析目录下的所有 Excel 文件"""
    parser = ExcelParser(
        output_dir=output_dir,
        include_hidden=include_hidden,
        include_empty=include_empty,
    )
    return parser.parse_directory(dir_path, recursive=recursive)
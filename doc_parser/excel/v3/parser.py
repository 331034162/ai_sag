"""
Excel V3 JSON 解析器
====================

提供 ExcelParser 类和便捷函数，将 Excel 文件解析为结构化的 JSON 数据。

用法：
    # 便捷函数
    from ai_sag.doc_parser.excel.v3 import parse_excel, parse_directory

    result = parse_excel("/path/to/file.xlsx")
    print(result.to_json())

    # 类实例
    from ai_sag.doc_parser.excel.v3 import ExcelParser

    parser = ExcelParser(output_dir="./output", signing_detection=True)
    result = parser.parse("/path/to/file.xlsx")
"""

import os
import logging
from pathlib import Path
from typing import Optional

import openpyxl

from ai_sag.doc_parser.excel.v3.config import ENABLE_SIGNING_DETECTION, INCLUDE_HIDDEN
from ai_sag.doc_parser.excel.v3.models import SheetJSON, ExcelJSON
from ai_sag.doc_parser.excel.v3.converter import sheet_to_json

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Excel 文档 JSON 解析器。

    将 Excel 文件解析为结构化 JSON 数据，保留合并单元格、货币格式、
    分组表头展平、表单字段、签章信息、批注、公式等。

    Args:
        output_dir: 输出目录，传值则保存 JSON 结果到磁盘，传 None 则仅返回内存结果
        signing_detection: 是否启用签章行检测
        include_hidden: 是否包含隐藏行列
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        signing_detection: bool = ENABLE_SIGNING_DETECTION,
        include_hidden: bool = INCLUDE_HIDDEN,
    ):
        self.output_dir = output_dir
        self.signing_detection = signing_detection
        self.include_hidden = include_hidden

    def parse(self, file_path: str) -> ExcelJSON:
        """
        解析单个 Excel 文件。

        Args:
            file_path: Excel 文件路径（.xlsx / .xls）

        Returns:
            ExcelJSON 解析结果
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx / .xls")

        # 动态设置配置
        from ai_sag.doc_parser.excel.v3 import config as cfg
        original_signing = cfg.ENABLE_SIGNING_DETECTION
        original_hidden = cfg.INCLUDE_HIDDEN
        cfg.ENABLE_SIGNING_DETECTION = self.signing_detection
        cfg.INCLUDE_HIDDEN = self.include_hidden

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            wb_formula = openpyxl.load_workbook(file_path, data_only=False)

            # 逐 Sheet 解析
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_f = wb_formula[sheet_name]
                sheet_data = sheet_to_json(
                    ws, sheet_title=sheet_name,
                    ws_formula=ws_f, include_hidden=self.include_hidden,
                )

                sc = SheetJSON(
                    sheet_name=sheet_data["sheet_name"],
                    title=sheet_data["title"],
                    sections=[
                        _make_section(s) for s in sheet_data["sections"]
                    ],
                    form_fields=sheet_data["form_fields"],
                    signing_info=sheet_data["signing_info"],
                    comments=sheet_data["comments"],
                    formulas=sheet_data["formulas"],
                    row_count=sheet_data["row_count"],
                    col_count=sheet_data["col_count"],
                )
                sheets.append(sc)

            result = ExcelJSON(
                file_path=str(path),
                file_name=path.name,
                total_sheets=len(sheets),
                sheets=sheets,
                metadata={
                    "signing_detection": self.signing_detection,
                    "include_hidden": self.include_hidden,
                    "sheet_names": wb.sheetnames,
                },
            )

            # 保存到磁盘
            if self.output_dir:
                from ai_sag.doc_parser.excel.v3.formatter import save_json
                out_dir = Path(self.output_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                stem = path.stem
                save_json(result, str(out_dir / f"{stem}.json"))

            return result

        finally:
            cfg.ENABLE_SIGNING_DETECTION = original_signing
            cfg.INCLUDE_HIDDEN = original_hidden

    def parse_directory(self, dir_path: str, recursive: bool = False) -> list[ExcelJSON]:
        """
        批量解析目录下的所有 Excel 文件。

        Args:
            dir_path: 目录路径
            recursive: 是否递归子目录

        Returns:
            ExcelJSON 列表
        """
        dp = Path(dir_path)
        if not dp.is_dir():
            raise NotADirectoryError(f"不是目录: {dir_path}")

        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        files = list(dp.glob(pattern))
        if not recursive:
            files += list(dp.glob("*.xls"))
        else:
            files += list(dp.glob("**/*.xls"))

        # 排除临时文件
        files = [f for f in files if not f.name.startswith("~$")]

        results = []
        for f in sorted(files):
            try:
                logger.info(f"解析: {f.name}")
                result = self.parse(str(f))
                results.append(result)
            except Exception as e:
                logger.error(f"解析失败 [{f.name}]: {e}")

        return results


# ============================================================
# 内部辅助
# ============================================================

def _make_section(data: dict) -> dict:
    """将 converter 输出的 section dict 规范化（确保 headers/rows 键存在）"""
    return {
        "headers": data.get("headers", []),
        "rows": data.get("rows", []),
    }


# ============================================================
# 便捷函数
# ============================================================

def parse_excel(
    file_path: str,
    output_dir: Optional[str] = None,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
    include_hidden: bool = INCLUDE_HIDDEN,
) -> ExcelJSON:
    """
    一键解析 Excel 文档为 JSON。

    Args:
        file_path: Excel 文件路径（.xlsx / .xls）
        output_dir: 输出目录，传值则保存 JSON 到磁盘，传 None 则仅返回内存结果
        signing_detection: 是否启用签章行检测
        include_hidden: 是否包含隐藏行列

    Returns:
        ExcelJSON 解析结果
    """
    parser = ExcelParser(
        output_dir=output_dir,
        signing_detection=signing_detection,
        include_hidden=include_hidden,
    )
    return parser.parse(file_path)


def parse_directory(
    dir_path: str,
    output_dir: Optional[str] = None,
    recursive: bool = False,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
    include_hidden: bool = INCLUDE_HIDDEN,
) -> list[ExcelJSON]:
    """
    一键批量解析目录下的所有 Excel 文件为 JSON。

    Args:
        dir_path: 目录路径
        output_dir: 输出目录
        recursive: 是否递归子目录
        signing_detection: 是否启用签章行检测
        include_hidden: 是否包含隐藏行列

    Returns:
        ExcelJSON 列表
    """
    parser = ExcelParser(
        output_dir=output_dir,
        signing_detection=signing_detection,
        include_hidden=include_hidden,
    )
    return parser.parse_directory(dir_path, recursive=recursive)

"""
Excel 解析器
============

提供 ExcelParser 类和便捷函数，将 Excel 文件解析为结构化的 ExcelResult。

用法：
    # 便捷函数
    from ai_sag.doc_parser.excel import parse_excel, parse_directory

    result = parse_excel("/path/to/file.xlsx")
    results = parse_directory("/path/to/dir")

    # 类实例
    from ai_sag.doc_parser.excel import ExcelParser

    parser = ExcelParser(output_dir="./output", signing_detection=True)
    result = parser.parse("/path/to/file.xlsx")
"""

import os
import logging
from pathlib import Path
from typing import Optional, Union

import openpyxl

from ai_sag.doc_parser.excel.v1.config import ENABLE_SIGNING_DETECTION
from ai_sag.doc_parser.excel.v1.models import SheetContent, ExcelResult
from ai_sag.doc_parser.excel.v1.converter import sheet_to_markdown

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Excel 文档解析器。

    将 Excel 文件解析为结构化结果，保留合并单元格、货币格式、
    分组表头展平、表单字段、签章信息等。

    Args:
        output_dir: 输出目录，传值则保存解析结果到磁盘，传 None 则仅返回内存结果
        signing_detection: 是否启用签章行检测，True=签章行剥离为元数据，False=当作普通数据行
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        signing_detection: bool = ENABLE_SIGNING_DETECTION,
    ):
        self.output_dir = output_dir
        self.signing_detection = signing_detection

    def parse(self, file_path: str) -> ExcelResult:
        """
        解析单个 Excel 文件。

        Args:
            file_path: Excel 文件路径（.xlsx / .xls）

        Returns:
            ExcelResult 解析结果
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的文件格式: {path.suffix}，仅支持 .xlsx / .xls")

        # 动态设置签章检测开关
        from ai_sag.doc_parser.excel.v1 import config as cfg
        original_signing = cfg.ENABLE_SIGNING_DETECTION
        cfg.ENABLE_SIGNING_DETECTION = self.signing_detection

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            # 构建完整 Markdown
            md_parts = []
            md_parts.append("# Excel 数据表")
            md_parts.append("")
            md_parts.append(f"> 源文件: `{file_path}`")
            md_parts.append(f"> 包含 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
            md_parts.append("")
            md_parts.append("## 目录")
            md_parts.append("")
            for idx, name in enumerate(wb.sheetnames, 1):
                md_parts.append(f"{idx}. [{name}](#{name.replace(' ', '-')})")
            md_parts.append("---")
            md_parts.append("")

            # 逐 Sheet 解析
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_md = sheet_to_markdown(ws, sheet_title=sheet_name)
                md_parts.append(sheet_md)

                # 从 Markdown 文本中提取表单/签章信息
                # （sheet_to_markdown 内部已处理，这里从 md 文本中简单提取）
                sc = SheetContent(
                    sheet_name=sheet_name,
                    markdown_text=sheet_md,
                    row_count=ws.max_row or 0,
                    col_count=ws.max_column or 0,
                )
                sheets.append(sc)

            full_markdown = "\n".join(md_parts)

            # 纯文本（去掉 Markdown 语法标记）
            full_text = _markdown_to_plain(full_markdown)

            # 从各 sheet 的 markdown 中提取元数据
            for sc in sheets:
                _extract_sheet_metadata(sc)

            result = ExcelResult(
                file_path=str(path),
                file_name=path.name,
                total_sheets=len(sheets),
                sheets=sheets,
                full_text=full_text,
                markdown_text=full_markdown,
                metadata={
                    "signing_detection": self.signing_detection,
                    "sheet_names": wb.sheetnames,
                },
            )

            # 保存到磁盘
            if self.output_dir:
                from ai_sag.doc_parser.excel.v1.formatter import save_markdown, save_text
                out_dir = Path(self.output_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                stem = path.stem
                save_markdown(result, str(out_dir / f"{stem}.md"))
                save_text(result, str(out_dir / f"{stem}.txt"))

            return result

        finally:
            # 恢复原始签章检测设置
            cfg.ENABLE_SIGNING_DETECTION = original_signing

    def parse_directory(self, dir_path: str, recursive: bool = False) -> list[ExcelResult]:
        """
        批量解析目录下的所有 Excel 文件。

        Args:
            dir_path: 目录路径
            recursive: 是否递归子目录

        Returns:
            ExcelResult 列表
        """
        dp = Path(dir_path)
        if not dp.is_dir():
            raise NotADirectoryError(f"不是目录: {dir_path}")

        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        files = list(dp.glob(pattern))
        if not recursive:
            files += list(dp.glob("*.xls"))
        else:
            files += list(dp.glob("**/*.xls"))

        # 排除临时文件
        files = [f for f in files if not f.name.startswith("~$")]

        results = []
        for f in sorted(files):
            try:
                logger.info(f"解析: {f.name}")
                result = self.parse(str(f))
                results.append(result)
            except Exception as e:
                logger.error(f"解析失败 [{f.name}]: {e}")

        return results


# ============================================================
# 内部辅助
# ============================================================

def _markdown_to_plain(md_text: str) -> str:
    """将 Markdown 转为纯文本（去掉表格竖线等格式标记）"""
    lines = md_text.split("\n")
    plain_lines = []
    for line in lines:
        stripped = line.strip()
        # 跳过分隔行
        if stripped.startswith("|") and all(c in "|- " for c in stripped):
            continue
        # 去掉表格竖线
        if stripped.startswith("|"):
            cells = [c.strip() for c in stripped.split("|") if c.strip()]
            plain_lines.append(" | ".join(cells))
        else:
            plain_lines.append(line)
    return "\n".join(plain_lines)


def _extract_sheet_metadata(sc: SheetContent):
    """从 SheetContent 的 markdown 文本中提取表单字段和签章信息"""
    sc.title = ""
    lines = sc.markdown_text.split("\n")
    in_form_section = False
    in_signing_section = False

    for line in lines:
        line_stripped = line.strip()

        # 标题（加粗文本）
        if line_stripped.startswith("**") and line_stripped.endswith("**"):
            label = line_stripped.strip("*").strip()
            if label.startswith("表单信息"):
                in_form_section = True
                in_signing_section = False
                continue
            elif label.startswith("签章信息"):
                in_signing_section = True
                in_form_section = False
                continue
            else:
                in_form_section = False
                in_signing_section = False
                if not sc.title:
                    sc.title = label

        # 列表项
        if line_stripped.startswith("- "):
            item = line_stripped[2:].strip()
            if in_form_section:
                sc.form_fields.append(item)
            elif in_signing_section:
                sc.signing_info.append(item)

        # 非列表项、非标题 → 退出 section
        if not line_stripped.startswith("- ") and not line_stripped.startswith("**") and line_stripped:
            if not line_stripped.startswith("|") and not line_stripped.startswith("#"):
                in_form_section = False
                in_signing_section = False


# ============================================================
# 便捷函数
# ============================================================

def parse_excel(
    file_path: str,
    output_dir: Optional[str] = None,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
) -> ExcelResult:
    """
    一键解析 Excel 文档。

    Args:
        file_path: Excel 文件路径（.xlsx / .xls）
        output_dir: 输出目录，传值则保存解析结果到磁盘，传 None 则仅返回内存结果
        signing_detection: 是否启用签章行检测

    Returns:
        ExcelResult 解析结果
    """
    parser = ExcelParser(output_dir=output_dir, signing_detection=signing_detection)
    return parser.parse(file_path)


def parse_directory(
    dir_path: str,
    output_dir: Optional[str] = None,
    recursive: bool = False,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
) -> list[ExcelResult]:
    """
    一键批量解析目录下的所有 Excel 文件。

    Args:
        dir_path: 目录路径
        output_dir: 输出目录
        recursive: 是否递归子目录
        signing_detection: 是否启用签章行检测

    Returns:
        ExcelResult 列表
    """
    parser = ExcelParser(output_dir=output_dir, signing_detection=signing_detection)
    return parser.parse_directory(dir_path, recursive=recursive)

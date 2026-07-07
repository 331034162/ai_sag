"""
Excel 单元格解析与 Markdown 转换核心逻辑
=========================================

处理流程：
1. 采集信息：预读所有合并单元格信息 + 每个单元格的值和元数据
2. 格式化值：根据元数据（number_format等）格式化每个单元格的值
3. 合并处理：按合并规则处理每个单元格的最终输出值
4. 结构检测：自动识别标题行/表头行/表单行/签章行/多段表格
5. Markdown输出：生成语义完整的Markdown表格
"""

import re
from dataclasses import dataclass, field
from typing import Optional, Any, List, Dict, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

from .config import ENABLE_SIGNING_DETECTION, SIGNING_KEYWORDS


# ============================================================
# 数据结构
# ============================================================

@dataclass
class MergeInfo:
    """一个合并区域的完整信息"""
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def row_span(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def col_span(self) -> int:
        return self.max_col - self.min_col + 1

    def contains(self, row: int, col: int) -> bool:
        return self.min_row <= row <= self.max_row and self.min_col <= col <= self.max_col

    def is_origin(self, row: int, col: int) -> bool:
        return row == self.min_row and col == self.min_col


@dataclass
class CellMeta:
    """单个单元格的完整信息"""
    raw_value: Any = None
    number_format: str = "General"
    data_type: str = "s"         # s=string, n=number, b=bool, d=datetime
    formatted: str = ""          # 根据元数据格式化后的值
    final: str = ""              # 合并处理后的最终输出值
    merge: Optional[MergeInfo] = None
    is_merge_origin: bool = False
    col_span: int = 1            # 作为合并起点时的列跨度
    row_span: int = 1            # 作为合并起点时的行跨度


# ============================================================
# 阶段1：信息采集
# ============================================================

def collect_merge_info(ws) -> Tuple[Dict[Tuple, MergeInfo], List[MergeInfo]]:
    """
    采集所有合并单元格信息。

    Returns:
        cell_merge_map: {(row, col): MergeInfo} 每个单元格属于哪个合并区域
        merge_list: [MergeInfo] 所有合并区域列表
    """
    cell_merge_map: Dict[Tuple[int, int], MergeInfo] = {}
    merge_list: List[MergeInfo] = []

    for merged_range in ws.merged_cells.ranges:
        mi = MergeInfo(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        )
        merge_list.append(mi)
        for r in range(mi.min_row, mi.max_row + 1):
            for c in range(mi.min_col, mi.max_col + 1):
                cell_merge_map[(r, c)] = mi

    return cell_merge_map, merge_list


def collect_cell_meta(ws, cell_merge_map) -> Dict[Tuple[int, int], CellMeta]:
    """
    采集每个单元格的值和元数据。

    Returns:
        cell_map: {(row, col): CellMeta}
    """
    cell_map: Dict[Tuple[int, int], CellMeta] = {}

    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            merge = cell_merge_map.get((row_idx, col_idx))
            is_origin = merge is not None and merge.is_origin(row_idx, col_idx)

            cm = CellMeta(
                raw_value=cell.value,
                number_format=cell.number_format or "General",
                data_type=cell.data_type or "s",
                merge=merge,
                is_merge_origin=is_origin,
                col_span=merge.col_span if is_origin else 1,
                row_span=merge.row_span if is_origin else 1,
            )
            cell_map[(row_idx, col_idx)] = cm

    return cell_map


# ============================================================
# 阶段2：根据元数据格式化值
# ============================================================

def format_by_metadata(value: Any, number_format: str, data_type: str) -> str:
    """
    根据单元格的元数据格式化值。

    支持：货币(¥/$€)、百分比(%)、千分位、日期、文本
    """
    if value is None:
        return ""

    # 字符串
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>").strip()

    # 布尔
    if isinstance(value, bool):
        return "是" if value else "否"

    # 日期/时间
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)

    # 数值
    if isinstance(value, (int, float)):
        fmt = number_format or ""

        # 货币格式 — 始终保留2位小数
        for symbol, prefix in [("¥", "¥"), ("￥", "¥"), ("$", "$"), ("€", "€")]:
            if symbol in fmt:
                return f"{prefix}{value:,.2f}"

        # 百分比格式
        if "%" in fmt:
            pct = value * 100
            if pct == int(pct):
                return f"{int(pct)}%"
            return f"{pct:.2f}%"

        # 千分位格式
        if "#,##0" in fmt or "#,#0" in fmt:
            if ".00" in fmt or ".0" in fmt:
                return f"{value:,.2f}"
            if isinstance(value, float) and value != int(value):
                return f"{value:,.2f}"
            return f"{int(value):,}"

        # 通用数值
        if isinstance(value, float):
            if value == int(value):
                return str(int(value))
            return str(round(value, 10))

        return str(value)

    return str(value)


def format_all_cells(cell_map: Dict[Tuple[int, int], CellMeta]):
    """格式化所有单元格"""
    for cm in cell_map.values():
        cm.formatted = format_by_metadata(cm.raw_value, cm.number_format, cm.data_type)


# ============================================================
# 阶段3：按合并规则处理最终值
# ============================================================

def apply_merge_logic(cell_map: Dict[Tuple[int, int], CellMeta], merge_list: List[MergeInfo]):
    """
    根据合并信息处理每个单元格的最终输出值。

    规则：
    - 非合并单元格：直接使用格式化值
    - 合并左上角：保留格式化值
    - 行方向合并（同列不同行）：填充左上角的值 → 保证每行语义完整
    - 列方向合并（同行不同列）：左上角保留值，其余留空
    - 块合并（行+列）：每行同列位置填值，其余列留空
    """
    origin_values: Dict[Tuple[int, int], str] = {}
    for mi in merge_list:
        origin = cell_map.get((mi.min_row, mi.min_col))
        if origin:
            origin_values[(mi.min_row, mi.min_col)] = origin.formatted

    for (row, col), cm in cell_map.items():
        if cm.merge is None or cm.is_merge_origin:
            cm.final = cm.formatted
        else:
            mi = cm.merge
            origin_val = origin_values.get((mi.min_row, mi.min_col), "")
            if col == mi.min_col:
                cm.final = origin_val
            else:
                cm.final = ""


# ============================================================
# 阶段4：构建行数据 + 结构检测
# ============================================================

def build_rows(cell_map: Dict[Tuple[int, int], CellMeta], num_rows: int, num_cols: int):
    """构建行数据列表"""
    rows = []
    for row_idx in range(1, num_rows + 1):
        row_data = []
        for col_idx in range(1, num_cols + 1):
            cm = cell_map.get((row_idx, col_idx))
            row_data.append(cm.final if cm else "")
        rows.append((row_idx, row_data))
    return rows


def is_empty_row(row_data: list) -> bool:
    return all(v.strip() == "" for v in row_data)


def _has_col_merges(row_idx: int, cell_map: Dict, num_cols: int) -> bool:
    """判断该行是否有列合并起点"""
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span > 1:
            return True
    return False


def _is_title_row(row_idx: int, row_data: list, cell_map: Dict, num_cols: int) -> bool:
    """标题行：单个值横跨大部分列"""
    non_empty = [v for v in row_data if v.strip()]
    if len(non_empty) != 1:
        return False
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span >= num_cols * 0.5:
            return True
    return False


def _is_form_row(row_idx: int, cell_map: Dict, num_cols: int) -> bool:
    """表单字段行：合并起点含 key：value"""
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span > 1:
            val = cm.formatted
            if "：" in val or ":" in val:
                return True
    return False


def _is_signing_row(row_data: list) -> bool:
    """
    签章行检测。

    严格匹配：签字/签章/公章等关键词直接匹配。
    审批关键词：仅匹配标签形式（如"部门经理审批："），不匹配描述文本中的嵌入词。
    """
    for v in row_data:
        v = v.strip()
        if not v:
            continue
        if any(kw in v for kw in SIGNING_KEYWORDS):
            return True
        if '审批' in v:
            if re.search(r'审批[：:]', v):
                return True
            if len(v) <= 10 and v.endswith('审批'):
                return True
    return False


def _is_group_header(row_idx: int, rows: list, cell_map: Dict, num_cols: int) -> bool:
    """
    判断是否为段落分组表头行。
    条件：
    1. 该行有列合并
    2. 后续紧跟着一行叶子表头（无列合并、非空值够多、多为文本）
    """
    if not _has_col_merges(row_idx, cell_map, num_cols):
        return False

    row_pos = None
    for i, (ri, _) in enumerate(rows):
        if ri == row_idx:
            row_pos = i
            break
    if row_pos is None:
        return False

    for j in range(row_pos + 1, min(row_pos + 5, len(rows))):
        _, next_data = rows[j]
        if is_empty_row(next_data):
            continue
        next_row_idx = rows[j][0]
        next_has_merges = _has_col_merges(next_row_idx, cell_map, num_cols)
        if next_has_merges:
            return True
        non_empty = sum(1 for v in next_data if v.strip())
        if non_empty >= max(num_cols * 0.3, 2):
            return True
        break

    return False


def _extract_form_fields(row_idx: int, cell_map: Dict, num_cols: int) -> List[str]:
    """
    从表单行提取 key:value 字段，组合标签和值。

    例: A10:B10="申请人签字：" + C10:D10="张三" → "申请人签字：张三"
    """
    merges = []
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span > 1:
            merges.append((c, c + cm.col_span - 1, cm.formatted.strip()))

    fields = []
    for i, (start, end, val) in enumerate(merges):
        if not val or ("：" not in val and ":" not in val):
            continue
        next_val = ""
        for j in range(i + 1, len(merges)):
            ns, ne, nv = merges[j]
            if ns == end + 1:
                next_val = nv.replace("<br>", "")
                break
        if not next_val:
            for c in range(end + 1, num_cols + 1):
                cm = cell_map.get((row_idx, c))
                if cm and cm.formatted.strip():
                    next_val = cm.formatted.strip().replace("<br>", "")
                    break
        if next_val:
            fields.append(f"{val}{next_val}")
        else:
            fields.append(val)
    return fields


def _extract_signing_info(row_data: list) -> List[str]:
    """
    从签章行提取信息，组合 label:value。

    例: ["申请人签字：", "", "张三", "", "企业公章签章区域"]
    → ["申请人签字：张三", "企业公章签章区域"]
    """
    parts = []
    for v in row_data:
        v = v.strip().replace("<br>", "")
        if not v or v == "签章区域":
            continue
        parts.append(v)

    results = []
    i = 0
    while i < len(parts):
        v = parts[i]
        if any(kw in v for kw in ['签章', '公章', '财务专用章', '法人名章']):
            results.append(v)
            i += 1
            continue
        if '：' in v or ':' in v:
            if i + 1 < len(parts) and '：' not in parts[i + 1] and ':' not in parts[i + 1]:
                results.append(f"{v}{parts[i + 1]}")
                i += 2
                continue
        results.append(v)
        i += 1

    return results


def _flatten_header(group_rows: List[Tuple[int, list]], leaf_data: list,
                    cell_map: Dict, num_cols: int,
                    first_col_value: str = None) -> list:
    """
    将多行表头展平为单行。

    例: "基础工商信息" + "公司名称" → "基础工商信息/公司名称"
    """
    flattened = list(leaf_data)

    if first_col_value is not None and flattened[0].strip() == "":
        flattened[0] = first_col_value

    if not group_rows:
        return flattened

    for col_idx in range(num_cols):
        col_num = col_idx + 1
        leaf_name = flattened[col_idx].strip() if col_idx < len(flattened) else ""

        parts = []
        if leaf_name:
            parts.append(leaf_name)

        for g_row_idx, _ in reversed(group_rows):
            group_name = ""
            for g in range(1, num_cols + 1):
                cm = cell_map.get((g_row_idx, g))
                if cm and cm.is_merge_origin and cm.col_span > 1:
                    if cm.merge and cm.merge.contains(g_row_idx, col_num):
                        group_name = cm.formatted.strip()
                        break
            if group_name and group_name != leaf_name and group_name not in parts:
                parts.insert(0, group_name)

        if parts:
            flattened[col_idx] = "/".join(parts)
        else:
            flattened[col_idx] = ""

    return flattened


# ============================================================
# 阶段5：Markdown 输出
# ============================================================

def _escape_markdown(text: str) -> str:
    """转义 Markdown 表格单元格中的特殊字符"""
    text = text.replace("|", "\\|")
    text = text.replace("*", "\\*")
    text = text.replace("_", "\\_")
    return text


def _output_table(lines: list, header: list, data_rows: list, num_cols: int):
    """输出一个 Markdown 表格"""
    while len(header) < num_cols:
        header.append("")
    for row in data_rows:
        while len(row) < num_cols:
            row.append("")

    # 裁剪尾部全空列
    while num_cols > 1:
        ci = num_cols - 1
        if header[ci].strip() == "" and all(r[ci].strip() == "" for r in data_rows):
            header.pop()
            for r in data_rows:
                r.pop()
            num_cols -= 1
        else:
            break

    # 空表头名用列字母填充
    for i, h in enumerate(header):
        if h.strip() == "":
            header[i] = f"列{get_column_letter(i + 1)}"

    if not data_rows and not any(v.strip() for v in header):
        return

    lines.append("| " + " | ".join(_escape_markdown(h) for h in header) + " |")
    lines.append("| " + " | ".join(["---"] * num_cols) + " |")
    for row in data_rows:
        lines.append("| " + " | ".join(_escape_markdown(v) for v in row) + " |")
    lines.append("")


# ============================================================
# 主处理函数
# ============================================================

def sheet_to_markdown(ws, sheet_title: Optional[str] = None) -> str:
    """
    将单个 Sheet 转换为 Markdown。

    处理流程：采集 → 格式化 → 合并处理 → 结构检测 → 输出
    """
    lines = []
    title = sheet_title or ws.title
    lines.append(f"## {title}")
    lines.append("")

    if ws.max_row is None or ws.max_column is None or ws.max_row == 0:
        lines.append("*（空表格）*")
        lines.append("")
        return "\n".join(lines)

    num_rows = ws.max_row
    num_cols = ws.max_column

    # ---- 阶段1：采集 ----
    cell_merge_map, merge_list = collect_merge_info(ws)
    cell_map = collect_cell_meta(ws, cell_merge_map)

    # ---- 阶段2：格式化 ----
    format_all_cells(cell_map)

    # ---- 阶段3：合并处理 ----
    apply_merge_logic(cell_map, merge_list)

    # ---- 阶段4：构建行数据 ----
    rows = build_rows(cell_map, num_rows, num_cols)

    # 找首列的行合并值
    first_col_merge_value = None
    for mi in merge_list:
        if mi.min_col == 1 and mi.row_span > 1:
            origin = cell_map.get((mi.min_row, mi.min_col))
            if origin:
                first_col_merge_value = origin.formatted
                break

    # ---- 结构检测与分段 ----
    title_texts = []
    form_metadata = []
    signing_metadata = []

    sections = []
    cur_groups = []
    cur_leaf = None
    cur_data = []

    def flush():
        if cur_leaf is not None:
            sections.append({
                "groups": list(cur_groups),
                "leaf": cur_leaf,
                "data": list(cur_data),
            })

    for pos, (row_idx, row_data) in enumerate(rows):
        if is_empty_row(row_data):
            continue

        # 1. 标题行
        if _is_title_row(row_idx, row_data, cell_map, num_cols):
            flush()
            cur_groups, cur_leaf, cur_data = [], None, []
            for c in range(1, num_cols + 1):
                cm = cell_map.get((row_idx, c))
                if cm and cm.is_merge_origin and cm.col_span >= num_cols * 0.5:
                    title_texts.append(cm.formatted)
                    break
            else:
                title_texts.append(next((v for v in row_data if v.strip()), ""))
            continue

        # 2. 签章行
        if ENABLE_SIGNING_DETECTION and _is_signing_row(row_data):
            flush()
            cur_groups, cur_leaf, cur_data = [], None, []
            for s in _extract_signing_info(row_data):
                if s and s not in signing_metadata:
                    signing_metadata.append(s)
            continue

        # 3. 表单字段行
        if _is_form_row(row_idx, cell_map, num_cols):
            flush()
            cur_groups, cur_leaf, cur_data = [], None, []
            for f in _extract_form_fields(row_idx, cell_map, num_cols):
                form_metadata.append(f)
            continue

        # 4. 段落分组表头行
        if _is_group_header(row_idx, rows, cell_map, num_cols):
            if cur_leaf is not None and cur_data:
                flush()
                cur_groups = []
                cur_data = []
            cur_groups.append(pos)
            cur_leaf = None
            continue

        # 6. 叶子表头行 vs 数据行
        has_merges = _has_col_merges(row_idx, cell_map, num_cols)
        non_empty = sum(1 for v in row_data if v.strip())
        numeric_count = sum(
            1 for v in row_data
            if v.strip() and re.match(r"^[\d,.¥￥$€%]+$", v.strip())
        )

        if not has_merges and non_empty >= max(num_cols * 0.3, 2):
            if cur_leaf is None:
                if cur_groups:
                    cur_leaf = pos
                elif numeric_count > non_empty * 0.4:
                    cur_leaf = pos
                else:
                    cur_leaf = pos
            else:
                if not cur_data and numeric_count <= non_empty * 0.3:
                    cur_data.append(pos)
                else:
                    cur_data.append(pos)
        else:
            if cur_leaf is not None:
                cur_data.append(pos)
            elif non_empty >= 2:
                cur_data.append(pos)

    flush()

    # ---- 阶段5：输出 ----
    for t in title_texts:
        lines.append(f"**{t.strip()}**")
        lines.append("")

    if form_metadata:
        lines.append("**表单信息：**")
        for fm in form_metadata:
            lines.append(f"- {fm}")
        lines.append("")

    for sec in sections:
        group_row_indices = sec["groups"]
        leaf_pos = sec["leaf"]
        data_positions = sec["data"]

        group_rows = [(rows[p][0], rows[p][1]) for p in group_row_indices]
        leaf_data = rows[leaf_pos][1] if leaf_pos is not None else []
        data_rows = [rows[p][1] for p in data_positions]

        header = _flatten_header(
            group_rows, leaf_data, cell_map, num_cols,
            first_col_value=first_col_merge_value,
        )
        actual_cols = max(len(header), max((len(r) for r in data_rows), default=0))
        _output_table(lines, header, data_rows, actual_cols)

    if not sections:
        data_positions = []
        leaf_pos = None
        for pos, (row_idx, row_data) in enumerate(rows):
            if is_empty_row(row_data):
                continue
            rt = None
            if _is_title_row(row_idx, row_data, cell_map, num_cols):
                rt = "title"
            elif ENABLE_SIGNING_DETECTION and _is_signing_row(row_data):
                rt = "signing"
            elif _is_form_row(row_idx, cell_map, num_cols):
                rt = "form"

            if rt == "title":
                for c in range(1, num_cols + 1):
                    cm = cell_map.get((row_idx, c))
                    if cm and cm.is_merge_origin and cm.col_span >= num_cols * 0.5:
                        title_texts.append(cm.formatted)
                        break
                continue
            if rt == "form":
                for f in _extract_form_fields(row_idx, cell_map, num_cols):
                    form_metadata.append(f)
                continue
            if rt == "signing":
                for s in _extract_signing_info(row_data):
                    if s and s not in signing_metadata:
                        signing_metadata.append(s)
                continue

            if leaf_pos is None:
                leaf_pos = pos
            else:
                data_positions.append(pos)

        if leaf_pos is not None:
            leaf_data = rows[leaf_pos][1]
            data_rows = [rows[p][1] for p in data_positions]
            header = _flatten_header([], leaf_data, cell_map, num_cols,
                                     first_col_value=first_col_merge_value)
            actual_cols = max(len(header), max((len(r) for r in data_rows), default=0))
            _output_table(lines, header, data_rows, actual_cols)

    if signing_metadata:
        lines.append("**签章信息：**")
        for sm in signing_metadata:
            lines.append(f"- {sm}")
        lines.append("")

    if not any(line.startswith("|") for line in lines) and not title_texts:
        lines.append("*（空表格）*")
        lines.append("")

    return "\n".join(lines)

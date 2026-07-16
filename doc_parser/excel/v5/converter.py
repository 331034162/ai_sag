"""
Excel 单元格解析与二维数组转换核心逻辑
======================================

处理流程：
1. 采集信息：预读所有合并单元格信息 + 每个单元格的值和元数据
2. 格式化值：根据元数据（number_format等）格式化每个单元格的值（用于结构检测）
3. 合并处理：按合并规则处理每个单元格的最终输出值
4. 结构检测：自动识别标题行/表头行/表单行/签章行/多段表格
5. 二维数组输出：表头为第一行，数据行保留原始类型（int/float/str）
"""

import re
from dataclasses import dataclass
from typing import Optional, Any, List, Dict, Tuple

from openpyxl.utils import get_column_letter

from .config import ENABLE_SIGNING_DETECTION, SIGNING_KEYWORDS


# ============================================================
# 数据结构
# ============================================================

@dataclass
class MergeInfo:
    """一个合并区域的完整信息"""
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def row_span(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def col_span(self) -> int:
        return self.max_col - self.min_col + 1

    def contains(self, row: int, col: int) -> bool:
        return self.min_row <= row <= self.max_row and self.min_col <= col <= self.max_col

    def is_origin(self, row: int, col: int) -> bool:
        return row == self.min_row and col == self.min_col


@dataclass
class CellMeta:
    """单个单元格的完整信息"""
    raw_value: Any = None
    number_format: str = "General"
    data_type: str = "s"
    formatted: str = ""
    final: str = ""
    merge: Optional[MergeInfo] = None
    is_merge_origin: bool = False
    col_span: int = 1
    row_span: int = 1


# ============================================================
# 隐藏行列检测 & 公式采集
# ============================================================

def _has_active_autofilter(ws) -> bool:
    """检测 sheet 是否存在生效中的 AutoFilter 筛选条件。

    openpyxl 的 AutoFilter 陷阱：Excel 开启数据筛选后，不满足筛选条件的行会被
    标记为 row_dimensions[r].hidden=True。这种"隐藏"与用户手动右键隐藏的行
    在 hidden 属性上无法区分，但语义完全不同：
    - 手动隐藏：用户有意排除的辅助行/计算行，可过滤
    - AutoFilter 隐藏：用户在 Excel 里的临时查看筛选，数据本身完整，入库必须读取

    Excel 有两种筛选机制，均会引发上述 hidden 标记，需同时检测：
    1. sheet 级 AutoFilter：选中区域→数据→筛选，对应 ws.auto_filter
    2. Table 级筛选：Ctrl+T 转成表格后表头筛选，对应 ws.tables[*].autoFilter
       （注意 openpyxl 命名不一致：sheet 级是 auto_filter，Table 级是 autoFilter）

    判定依据：筛选区域非空（ref）且实际勾选了筛选值（filterColumn 非空）。
    仅 ref 没有 filterColumn 时是"未应用筛选条件"，不会产生隐藏行，无需特殊处理。
    """
    # 1. sheet 级 AutoFilter
    af = getattr(ws, "auto_filter", None)
    if af is not None and af.ref and af.filterColumn:
        return True
    # 2. Table 级筛选（ListObject）
    tables = getattr(ws, "tables", None)
    if tables:
        for tbl in tables.values():
            taf = getattr(tbl, "autoFilter", None)
            if taf is not None and taf.filterColumn:
                return True
    return False


def _get_visible_rows_cols(ws):
    """获取可见（非隐藏）的行号和列号列表。

    AutoFilter 修正：当 sheet 存在生效的 AutoFilter 筛选条件时，忽略所有行的
    hidden 标记——因为这些隐藏是筛选产生的，数据本身完整，入库场景必须全部读取。
    否则正常过滤手动隐藏的行列。
    """
    # AutoFilter 场景：筛选产生的隐藏行不算真隐藏，全部读取
    if _has_active_autofilter(ws):
        visible_rows = list(range(1, (ws.max_row or 0) + 1))
        visible_cols = list(range(1, (ws.max_column or 0) + 1))
        return visible_rows, visible_cols

    hidden_rows = set()
    for r, dim in ws.row_dimensions.items():
        if dim.hidden:
            hidden_rows.add(r)

    hidden_cols = set()
    for col_key, dim in ws.column_dimensions.items():
        if dim.hidden:
            if isinstance(col_key, int):
                hidden_cols.add(col_key)
            else:
                from openpyxl.utils import column_index_from_string
                hidden_cols.add(column_index_from_string(col_key))

    visible_rows = sorted(r for r in range(1, (ws.max_row or 0) + 1) if r not in hidden_rows)
    visible_cols = sorted(c for c in range(1, (ws.max_column or 0) + 1) if c not in hidden_cols)
    return visible_rows, visible_cols


def _collect_formulas(ws_formula, visible_rows, visible_cols) -> Dict[Tuple[int, int], str]:
    """采集公式文本: {(row, col): formula_text}"""
    formula_map: Dict[Tuple[int, int], str] = {}
    if ws_formula is None:
        return formula_map
    for row_idx in visible_rows:
        for col_idx in visible_cols:
            cell = ws_formula.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                formula_map[(row_idx, col_idx)] = cell.value
    return formula_map


# ============================================================
# 阶段1：信息采集
# ============================================================

def collect_merge_info(ws) -> Tuple[Dict[Tuple, MergeInfo], List[MergeInfo]]:
    cell_merge_map: Dict[Tuple[int, int], MergeInfo] = {}
    merge_list: List[MergeInfo] = []

    for merged_range in ws.merged_cells.ranges:
        mi = MergeInfo(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        )
        merge_list.append(mi)
        for r in range(mi.min_row, mi.max_row + 1):
            for c in range(mi.min_col, mi.max_col + 1):
                cell_merge_map[(r, c)] = mi

    return cell_merge_map, merge_list


def collect_cell_meta(ws, cell_merge_map, visible_rows, visible_cols) -> Dict[Tuple[int, int], CellMeta]:
    cell_map: Dict[Tuple[int, int], CellMeta] = {}

    for row_idx in visible_rows:
        for col_idx in visible_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            merge = cell_merge_map.get((row_idx, col_idx))
            is_origin = merge is not None and merge.is_origin(row_idx, col_idx)

            cm = CellMeta(
                raw_value=cell.value,
                number_format=cell.number_format or "General",
                data_type=cell.data_type or "s",
                merge=merge,
                is_merge_origin=is_origin,
                col_span=merge.col_span if is_origin else 1,
                row_span=merge.row_span if is_origin else 1,
            )
            cell_map[(row_idx, col_idx)] = cm

    return cell_map


# ============================================================
# 阶段2：格式化（用于结构检测）
# ============================================================

def format_by_metadata(value: Any, number_format: str, data_type: str) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>").strip()
    if isinstance(value, bool):
        return "是" if value else "否"
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    if isinstance(value, (int, float)):
        fmt = number_format or ""
        for symbol, prefix in [("¥", "¥"), ("￥", "¥"), ("$", "$"), ("€", "€")]:
            if symbol in fmt:
                return f"{prefix}{value:,.2f}"
        if "%" in fmt:
            pct = value * 100
            return f"{int(pct)}%" if pct == int(pct) else f"{pct:.2f}%"
        if "#,##0" in fmt or "#,#0" in fmt:
            if ".00" in fmt or ".0" in fmt:
                return f"{value:,.2f}"
            if isinstance(value, float) and value != int(value):
                return f"{value:,.2f}"
            return f"{int(value):,}"
        if isinstance(value, float):
            return str(int(value)) if value == int(value) else str(round(value, 10))
        return str(value)
    return str(value)


def format_all_cells(cell_map: Dict[Tuple[int, int], CellMeta]):
    for cm in cell_map.values():
        cm.formatted = format_by_metadata(cm.raw_value, cm.number_format, cm.data_type)


# ============================================================
# 阶段3：合并处理
# ============================================================

def apply_merge_logic(cell_map: Dict[Tuple[int, int], CellMeta], merge_list: List[MergeInfo]):
    origin_values: Dict[Tuple[int, int], str] = {}
    for mi in merge_list:
        origin = cell_map.get((mi.min_row, mi.min_col))
        if origin:
            origin_values[(mi.min_row, mi.min_col)] = origin.formatted

    for (row, col), cm in cell_map.items():
        if cm.merge is None or cm.is_merge_origin:
            cm.final = cm.formatted
        else:
            mi = cm.merge
            origin_val = origin_values.get((mi.min_row, mi.min_col), "")
            if col == mi.min_col:
                cm.final = origin_val
            else:
                cm.final = ""


# ============================================================
# 阶段4：构建行数据 + 结构检测
# ============================================================

def build_rows(cell_map, visible_rows, visible_cols):
    rows = []
    for row_idx in visible_rows:
        row_data = []
        for col_idx in visible_cols:
            cm = cell_map.get((row_idx, col_idx))
            row_data.append(cm.final if cm else "")
        rows.append((row_idx, row_data))
    return rows


def is_empty_row(row_data: list) -> bool:
    return all(v.strip() == "" for v in row_data)


def _has_col_merges(row_idx, cell_map, num_cols):
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span > 1:
            return True
    return False


def _is_title_row(row_idx, row_data, cell_map, num_cols):
    non_empty = [v for v in row_data if v.strip()]
    if len(non_empty) != 1:
        return False
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span >= num_cols * 0.5:
            return True
    return False


def _is_form_row(row_idx, cell_map, num_cols):
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span > 1:
            if "：" in cm.formatted or ":" in cm.formatted:
                return True
    return False


def _is_signing_row(row_data):
    for v in row_data:
        v = v.strip()
        if not v:
            continue
        if any(kw in v for kw in SIGNING_KEYWORDS):
            return True
        if '审批' in v:
            if re.search(r'审批[：:]', v):
                return True
            if len(v) <= 10 and v.endswith('审批'):
                return True
    return False


def _is_group_header(row_idx, rows, cell_map, num_cols):
    if not _has_col_merges(row_idx, cell_map, num_cols):
        return False
    row_pos = None
    for i, (ri, _) in enumerate(rows):
        if ri == row_idx:
            row_pos = i
            break
    if row_pos is None:
        return False
    for j in range(row_pos + 1, min(row_pos + 5, len(rows))):
        _, next_data = rows[j]
        if is_empty_row(next_data):
            continue
        next_row_idx = rows[j][0]
        if _has_col_merges(next_row_idx, cell_map, num_cols):
            return True
        non_empty = sum(1 for v in next_data if v.strip())
        if non_empty >= max(num_cols * 0.3, 2):
            return True
        break
    return False


def _extract_form_fields(row_idx, cell_map, num_cols):
    merges = []
    for c in range(1, num_cols + 1):
        cm = cell_map.get((row_idx, c))
        if cm and cm.is_merge_origin and cm.col_span > 1:
            merges.append((c, c + cm.col_span - 1, cm.formatted.strip()))
    fields = []
    for i, (start, end, val) in enumerate(merges):
        if not val or ("：" not in val and ":" not in val):
            continue
        next_val = ""
        for j in range(i + 1, len(merges)):
            ns, ne, nv = merges[j]
            if ns == end + 1:
                next_val = nv.replace("<br>", "")
                break
        if not next_val:
            for c in range(end + 1, num_cols + 1):
                cm = cell_map.get((row_idx, c))
                if cm and cm.formatted.strip():
                    next_val = cm.formatted.strip().replace("<br>", "")
                    break
        fields.append(f"{val}{next_val}" if next_val else val)
    return fields


def _extract_signing_info(row_data):
    parts = [v.strip().replace("<br>", "") for v in row_data if v.strip() and v.strip() != "签章区域"]
    results = []
    i = 0
    while i < len(parts):
        v = parts[i]
        if any(kw in v for kw in ['签章', '公章', '财务专用章', '法人名章']):
            results.append(v)
            i += 1
            continue
        if ('：' in v or ':' in v) and i + 1 < len(parts):
            if '：' not in parts[i + 1] and ':' not in parts[i + 1]:
                results.append(f"{v}{parts[i + 1]}")
                i += 2
                continue
        results.append(v)
        i += 1
    return results


def _flatten_header(group_rows, leaf_data, cell_map, num_cols,
                    visible_cols=None, first_col_value=None):
    flattened = list(leaf_data)
    if first_col_value is not None and flattened[0].strip() == "":
        flattened[0] = first_col_value

    if not group_rows:
        flattened = _filter_and_label_headers(flattened, visible_cols, num_cols)
        return flattened

    for col_idx in range(num_cols):
        col_num = col_idx + 1
        leaf_name = flattened[col_idx].strip() if col_idx < len(flattened) else ""
        parts = [leaf_name] if leaf_name else []

        for g_row_idx, _ in reversed(group_rows):
            group_name = ""
            for g in range(1, num_cols + 1):
                cm = cell_map.get((g_row_idx, g))
                if cm and cm.is_merge_origin and cm.col_span > 1:
                    if cm.merge and cm.merge.contains(g_row_idx, col_num):
                        group_name = cm.formatted.strip()
                        break
            if group_name and group_name != leaf_name and group_name not in parts:
                parts.insert(0, group_name)

        flattened[col_idx] = "/".join(parts) if parts else ""

    flattened = _filter_and_label_headers(flattened, visible_cols, num_cols)
    return flattened


def _filter_and_label_headers(headers, visible_cols, num_cols):
    if visible_cols is not None:
        indices = [c - 1 for c in visible_cols if 0 < c <= num_cols]
        headers = [headers[i] if i < len(headers) else "" for i in indices]
    for i, h in enumerate(headers):
        if h.strip() == "":
            col_idx = visible_cols[i] if visible_cols and i < len(visible_cols) else i + 1
            headers[i] = f"列{get_column_letter(col_idx)}"
    return headers


# ============================================================
# 原始类型值提取
# ============================================================

def _native_value(raw_value: Any) -> Any:
    """
    提取单元格的原始类型值，用于 JSON 输出。

    - int/float → 保留数值类型
    - 日期 → "YYYY-MM-DD" 字符串
    - 布尔 → "是"/"否"
    - 字符串 → 清理换行
    - None → None
    """
    if raw_value is None:
        return None
    if isinstance(raw_value, bool):
        return "是" if raw_value else "否"
    if isinstance(raw_value, (int, float)):
        return raw_value
    if hasattr(raw_value, "strftime"):
        try:
            return raw_value.strftime("%Y-%m-%d")
        except Exception:
            return str(raw_value)
    if isinstance(raw_value, str):
        return raw_value.replace("\r\n", "\n").replace("\r", "\n").strip()
    return str(raw_value)


# ============================================================
# 主处理函数
# ============================================================

def sheet_to_data(ws, sheet_title: Optional[str] = None,
                  ws_formula=None, include_hidden: bool = False) -> dict:
    """
    将单个 Sheet 转换为二维数组 JSON。

    数据行保留原始类型（int/float），公式同时记录文本和计算值。
    """
    title = sheet_title or ws.title

    empty_result = {
        "sheet_name": title, "title": "", "sections": [],
        "form_fields": [], "signing_info": [], "formulas": [],
        "row_count": 0, "col_count": 0,
    }

    if ws.max_row is None or ws.max_column is None or ws.max_row == 0:
        return empty_result

    if include_hidden:
        visible_rows = list(range(1, ws.max_row + 1))
        visible_cols = list(range(1, ws.max_column + 1))
    else:
        visible_rows, visible_cols = _get_visible_rows_cols(ws)

    if not visible_rows or not visible_cols:
        return empty_result

    num_cols = ws.max_column

    # 采集
    cell_merge_map, merge_list = collect_merge_info(ws)
    cell_map = collect_cell_meta(ws, cell_merge_map, visible_rows, visible_cols)
    formula_map = _collect_formulas(ws_formula, visible_rows, visible_cols)

    # 格式化 + 合并处理
    format_all_cells(cell_map)
    apply_merge_logic(cell_map, merge_list)

    # 构建行
    rows = build_rows(cell_map, visible_rows, visible_cols)

    # 首列行合并值
    first_col_merge_value = None
    for mi in merge_list:
        if mi.min_col == 1 and mi.row_span > 1:
            origin = cell_map.get((mi.min_row, mi.min_col))
            if origin:
                first_col_merge_value = origin.formatted
                break

    # 结构检测
    title_texts, form_metadata, signing_metadata = [], [], []
    sections = []
    cur_groups, cur_leaf, cur_data = [], None, []

    def flush():
        if cur_leaf is not None:
            sections.append({"groups": list(cur_groups), "leaf": cur_leaf, "data": list(cur_data)})

    for pos, (row_idx, row_data) in enumerate(rows):
        if is_empty_row(row_data):
            continue
        if _is_title_row(row_idx, row_data, cell_map, num_cols):
            flush()
            cur_groups, cur_leaf, cur_data = [], None, []
            for c in range(1, num_cols + 1):
                cm = cell_map.get((row_idx, c))
                if cm and cm.is_merge_origin and cm.col_span >= num_cols * 0.5:
                    title_texts.append(cm.formatted)
                    break
            else:
                title_texts.append(next((v for v in row_data if v.strip()), ""))
            continue
        if ENABLE_SIGNING_DETECTION and _is_signing_row(row_data):
            flush()
            cur_groups, cur_leaf, cur_data = [], None, []
            for s in _extract_signing_info(row_data):
                if s and s not in signing_metadata:
                    signing_metadata.append(s)
            continue
        if _is_form_row(row_idx, cell_map, num_cols):
            flush()
            cur_groups, cur_leaf, cur_data = [], None, []
            form_metadata.extend(_extract_form_fields(row_idx, cell_map, num_cols))
            continue
        if _is_group_header(row_idx, rows, cell_map, num_cols):
            if cur_leaf is not None and cur_data:
                flush()
                cur_groups, cur_data = [], []
            cur_groups.append(pos)
            cur_leaf = None
            continue
        has_merges = _has_col_merges(row_idx, cell_map, num_cols)
        non_empty = sum(1 for v in row_data if v.strip())
        if not has_merges and non_empty >= max(num_cols * 0.3, 2):
            if cur_leaf is None:
                cur_leaf = pos
            else:
                cur_data.append(pos)
        else:
            if cur_leaf is not None:
                cur_data.append(pos)
            elif non_empty >= 2:
                cur_data.append(pos)

    flush()

    # ---- 构建二维数组输出 ----
    sheet_title_text = title_texts[0] if title_texts else ""
    json_sections = []
    sections_data_positions = []

    # 可见列索引列表（用于过滤）
    vis_indices = [c - 1 for c in visible_cols if 0 < c <= num_cols] if visible_cols else None

    for sec in sections:
        leaf_pos = sec["leaf"]
        data_positions = sec["data"]

        group_rows = [(rows[p][0], rows[p][1]) for p in sec["groups"]]
        leaf_data = rows[leaf_pos][1] if leaf_pos is not None else []

        headers = _flatten_header(
            group_rows, leaf_data, cell_map, num_cols,
            visible_cols=visible_cols,
            first_col_value=first_col_merge_value,
        )

        # 数据行：原始类型值
        output_rows = []
        for pos in data_positions:
            if leaf_pos is not None and pos <= leaf_pos:
                continue
            row_idx = rows[pos][0]
            row_vals = []
            for vi, col_idx in enumerate(visible_cols):
                if vi >= len(headers):
                    break
                cm = cell_map.get((row_idx, col_idx))
                if cm and (row_idx, col_idx) in formula_map:
                    # 公式单元格：用计算值（可能为 None）
                    val = _native_value(cm.raw_value)
                elif cm:
                    val = _native_value(cm.raw_value)
                else:
                    val = None
                row_vals.append(val)
            # 裁剪到 headers 长度
            while len(row_vals) > len(headers):
                row_vals.pop()
            output_rows.append(row_vals)

        # 裁剪尾部空列
        headers, output_rows = _trim_empty_tail_cols(headers, output_rows)

        json_sections.append({"headers": headers, "rows": output_rows})
        sections_data_positions.append(data_positions)

    # Fallback
    if not json_sections:
        fb_headers, fb_rows, fb_positions = _fallback_extract(
            rows, cell_map, num_cols, visible_cols, first_col_merge_value,
            title_texts, form_metadata, signing_metadata, cell_map, formula_map,
        )
        if fb_headers:
            if title_texts and not sheet_title_text:
                sheet_title_text = title_texts[0]
            json_sections.append({"headers": fb_headers, "rows": fb_rows})
            sections_data_positions.append(fb_positions)

    # ---- 公式采集：关联到 section + row + header ----
    row_pos_map = _build_row_pos_map(sections, sections_data_positions, rows)
    all_formulas = _extract_formulas(
        cell_map, formula_map, row_pos_map, json_sections,
        visible_cols, num_cols,
    )

    total_rows = sum(len(s["rows"]) for s in json_sections)

    return {
        "sheet_name": title,
        "title": sheet_title_text,
        "sections": json_sections,
        "form_fields": form_metadata,
        "signing_info": signing_metadata,
        "formulas": all_formulas,
        "row_count": total_rows,
        "col_count": len(visible_cols),
    }


# ============================================================
# 辅助函数
# ============================================================

def _trim_empty_tail_cols(headers, data_rows):
    """裁剪尾部全空列（支持 None 值）"""
    num_cols = len(headers)
    while num_cols > 1:
        ci = num_cols - 1
        h_empty = headers[ci].strip() == ""
        all_empty = all(
            (r[ci] is None or (isinstance(r[ci], str) and r[ci].strip() == ""))
            if ci < len(r) else True
            for r in data_rows
        )
        if h_empty and all_empty:
            headers.pop()
            for r in data_rows:
                if ci < len(r):
                    r.pop()
            num_cols -= 1
        else:
            break
    return headers, data_rows


def _fallback_extract(rows, cell_map, num_cols, visible_cols, first_col_value,
                      title_texts, form_metadata, signing_metadata,
                      cell_map_full, formula_map):
    """Fallback：无段落结构时的简单表格提取"""
    leaf_pos = None
    data_positions = []

    for pos, (row_idx, row_data) in enumerate(rows):
        if is_empty_row(row_data):
            continue
        if _is_title_row(row_idx, row_data, cell_map_full, num_cols):
            for c in range(1, num_cols + 1):
                cm = cell_map_full.get((row_idx, c))
                if cm and cm.is_merge_origin and cm.col_span >= num_cols * 0.5:
                    title_texts.append(cm.formatted)
                    break
            continue
        if ENABLE_SIGNING_DETECTION and _is_signing_row(row_data):
            for s in _extract_signing_info(row_data):
                if s and s not in signing_metadata:
                    signing_metadata.append(s)
            continue
        if _is_form_row(row_idx, cell_map_full, num_cols):
            form_metadata.extend(_extract_form_fields(row_idx, cell_map_full, num_cols))
            continue
        if leaf_pos is None:
            leaf_pos = pos
        else:
            data_positions.append(pos)

    if leaf_pos is None:
        return None, [], []

    leaf_data = rows[leaf_pos][1]
    headers = _flatten_header([], leaf_data, cell_map_full, num_cols,
                              visible_cols=visible_cols, first_col_value=first_col_value)

    output_rows = []
    for pos in data_positions:
        row_idx = rows[pos][0]
        row_vals = []
        for vi, col_idx in enumerate(visible_cols):
            if vi >= len(headers):
                break
            cm = cell_map_full.get((row_idx, col_idx))
            row_vals.append(_native_value(cm.raw_value) if cm else None)
        while len(row_vals) > len(headers):
            row_vals.pop()
        output_rows.append(row_vals)

    headers, output_rows = _trim_empty_tail_cols(headers, output_rows)
    return headers, output_rows, data_positions


def _build_row_pos_map(sections, sections_data_positions, rows):
    """构建 行号 → (section_index, row_within_section) 映射"""
    row_pos_map: Dict[int, Tuple[int, int]] = {}
    for sec_idx, (sec, positions) in enumerate(zip(sections, sections_data_positions)):
        data_start = sec["leaf"] + 1 if sec["leaf"] is not None else 0
        for out_row, pos in enumerate(positions):
            if pos >= data_start:
                row_idx = rows[pos][0]
                row_pos_map[row_idx] = (sec_idx, out_row)
    return row_pos_map


def _extract_formulas(cell_map, formula_map, row_pos_map,
                      json_sections, visible_cols, num_cols) -> List[dict]:
    """提取公式，关联到 section/row/header，同时记录计算值"""
    from openpyxl.utils import column_index_from_string

    all_formulas = []
    for (row_idx, col_idx), formula_text in formula_map.items():
        cm = cell_map.get((row_idx, col_idx))
        computed = _native_value(cm.raw_value) if cm else None

        entry = {
            "cell": f"{get_column_letter(col_idx)}{row_idx}",
            "formula": formula_text,
            "value": computed,
        }

        # 关联到结构化位置
        if row_idx in row_pos_map:
            sec_idx, row_in_sec = row_pos_map[row_idx]
            entry["section"] = sec_idx
            entry["row"] = row_in_sec
            if visible_cols and col_idx in visible_cols:
                vi = visible_cols.index(col_idx)
                if vi < len(json_sections[sec_idx]["headers"]):
                    entry["header"] = json_sections[sec_idx]["headers"][vi]

        all_formulas.append(entry)

    return all_formulas
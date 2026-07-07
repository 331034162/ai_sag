"""
Excel V5 二维数组 JSON 解析器
==============================

提供 ExcelParser 类和便捷函数，将 Excel 解析为二维数组格式，
数据行保留原始类型（int/float），公式同时记录文本和计算值。

用法：
    from ai_sag.doc_parser.excel.v5 import parse_excel

    result = parse_excel("/path/to/file.xlsx")
    print(result.to_json())
"""

import logging
from pathlib import Path
from typing import Optional

import openpyxl

from ai_sag.doc_parser.excel.v5.config import ENABLE_SIGNING_DETECTION, INCLUDE_HIDDEN
from ai_sag.doc_parser.excel.v5.models import Section, SheetData, ExcelData
from ai_sag.doc_parser.excel.v5.converter import sheet_to_data

logger = logging.getLogger(__name__)


class ExcelParser:
    """
    Excel 二维数组 JSON 解析器。

    将 Excel 解析为二维数组格式：
    - 数据行保留原始类型（int/float）
    - 公式同时记录公式文本和计算值
    - 自动检测标题行、分组表头、表单字段、签章行

    Args:
        output_dir: 输出目录
        signing_detection: 是否启用签章行检测
        include_hidden: 是否包含隐藏行列
    """

    def __init__(
        self,
        output_dir: Optional[str] = None,
        signing_detection: bool = ENABLE_SIGNING_DETECTION,
        include_hidden: bool = INCLUDE_HIDDEN,
    ):
        self.output_dir = output_dir
        self.signing_detection = signing_detection
        self.include_hidden = include_hidden

    def parse(self, file_path: str) -> ExcelData:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        if path.suffix.lower() not in (".xlsx", ".xls"):
            raise ValueError(f"不支持的文件格式: {path.suffix}")

        from ai_sag.doc_parser.excel.v5 import config as cfg
        original_signing = cfg.ENABLE_SIGNING_DETECTION
        original_hidden = cfg.INCLUDE_HIDDEN
        cfg.ENABLE_SIGNING_DETECTION = self.signing_detection
        cfg.INCLUDE_HIDDEN = self.include_hidden

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            wb_formula = openpyxl.load_workbook(file_path, data_only=False)

            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws_f = wb_formula[sheet_name]
                data = sheet_to_data(
                    ws, sheet_title=sheet_name,
                    ws_formula=ws_f, include_hidden=self.include_hidden,
                )

                sc = SheetData(
                    sheet_name=data["sheet_name"],
                    title=data["title"],
                    sections=[Section(headers=s["headers"], rows=s["rows"])
                              for s in data["sections"]],
                    form_fields=data["form_fields"],
                    signing_info=data["signing_info"],
                    formulas=data["formulas"],
                    row_count=data["row_count"],
                    col_count=data["col_count"],
                )
                sheets.append(sc)

            result = ExcelData(
                file_path=str(path),
                file_name=path.name,
                total_sheets=len(sheets),
                sheets=sheets,
                metadata={
                    "signing_detection": self.signing_detection,
                    "include_hidden": self.include_hidden,
                    "sheet_names": wb.sheetnames,
                },
            )

            if self.output_dir:
                from ai_sag.doc_parser.excel.v5.formatter import save_json
                out_dir = Path(self.output_dir)
                out_dir.mkdir(parents=True, exist_ok=True)
                save_json(result, str(out_dir / f"{path.stem}.json"))

            return result

        finally:
            cfg.ENABLE_SIGNING_DETECTION = original_signing
            cfg.INCLUDE_HIDDEN = original_hidden

    def parse_directory(self, dir_path: str, recursive: bool = False) -> list[ExcelData]:
        dp = Path(dir_path)
        if not dp.is_dir():
            raise NotADirectoryError(f"不是目录: {dir_path}")

        pattern = "**/*.xlsx" if recursive else "*.xlsx"
        files = list(dp.glob(pattern))
        if not recursive:
            files += list(dp.glob("*.xls"))
        else:
            files += list(dp.glob("**/*.xls"))

        files = [f for f in files if not f.name.startswith("~$")]

        results = []
        for f in sorted(files):
            try:
                logger.info(f"解析: {f.name}")
                results.append(self.parse(str(f)))
            except Exception as e:
                logger.error(f"解析失败 [{f.name}]: {e}")

        return results


def parse_excel(
    file_path: str,
    output_dir: Optional[str] = None,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
    include_hidden: bool = INCLUDE_HIDDEN,
) -> ExcelData:
    parser = ExcelParser(
        output_dir=output_dir,
        signing_detection=signing_detection,
        include_hidden=include_hidden,
    )
    return parser.parse(file_path)


def parse_directory(
    dir_path: str,
    output_dir: Optional[str] = None,
    recursive: bool = False,
    signing_detection: bool = ENABLE_SIGNING_DETECTION,
    include_hidden: bool = INCLUDE_HIDDEN,
) -> list[ExcelData]:
    parser = ExcelParser(
        output_dir=output_dir,
        signing_detection=signing_detection,
        include_hidden=include_hidden,
    )
    return parser.parse_directory(dir_path, recursive=recursive)
